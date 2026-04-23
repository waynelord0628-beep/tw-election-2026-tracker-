"""
將鏈上調查結果（代幣來源）補充進 Excel
新增一個「代幣來源調查」工作表
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_PATH = r"E:\polymarket選舉賭博\四地址台灣選舉交易.xlsx"

thin = Side(style="thin", color="CCCCCC")


def border():
    return Border(top=thin, left=thin, right=thin, bottom=thin)


findings = [
    # (地址, 名稱, 市場, 動作, 日期, 數量, 價格, 金額USDC, 交易Hash, 說明)
    (
        "0x426227d4a9c4ad5a1aae7f2706238f2154b9abaa",
        "Kuomintang",
        "KMT",
        "BUY（鏈上直接合約）",
        "2025-12-26 21:47:31",
        100,
        0.84,
        84.0,
        "0x23345629edb3f8ca6555eff4f1c88cd35ae83ec686d047132792768506161bd4",
        "透過 NegRiskCtfExchange matchOrders 直接買入，Polymarket trades API 未收錄此筆",
    ),
    (
        "0x426227d4a9c4ad5a1aae7f2706238f2154b9abaa",
        "Kuomintang",
        "KMT",
        "SELL（Polymarket記錄）",
        "2026-02-07 01:20:11",
        100,
        0.87,
        87.0,
        "0x3d3c6abf90de71efc9775134c58afaf6f55c8d18d6e1fdfc8aa03f1bb380ad1a",
        "賣出先前買入的100股，獲利約 $3（$87-$84=+$3）",
    ),
    (
        "0x0d5ee5c536ccd78d179487b3d7e43ae4304d5c24",
        "ChiangWan-an",
        "KMT",
        "BUY？（來源不明）",
        "不明",
        100,
        "?",
        "?",
        "無",
        "推測透過 negRisk adapter 持倉（買入其他候選人 No ≈ 持有 KMT Yes），此操作不產生 ERC-1155 直接 Transfer；或為 USDC→Yes+No split 操作",
    ),
    (
        "0x0d5ee5c536ccd78d179487b3d7e43ae4304d5c24",
        "ChiangWan-an",
        "KMT",
        "SELL（Polymarket記錄）",
        "2026-02-07 01:20:33",
        20,
        0.87,
        17.4,
        "0xbaa910668e8413051767f452e0750debe9d7f03bd578681bbccbdbd3270398e0",
        "賣出20股",
    ),
    (
        "0x0d5ee5c536ccd78d179487b3d7e43ae4304d5c24",
        "ChiangWan-an",
        "KMT",
        "SELL（Polymarket記錄）",
        "2026-04-10 12:37:13",
        80,
        0.86,
        68.8,
        "0x789c588349aadb4724642ee165a25d4dd79231909a03d757cd2c7edb471fa63c",
        "鏈上確認：ChiangWan-an 直接轉出80個KMT-Yes至NegRiskExchange，收到 68.8 USDC",
    ),
]

wb = openpyxl.load_workbook(EXCEL_PATH)

# 移除舊工作表（若存在）
for name in ["代幣來源調查"]:
    if name in wb.sheetnames:
        del wb[name]

ws = wb.create_sheet("代幣來源調查")

headers = [
    "錢包地址",
    "名稱",
    "市場",
    "操作類型",
    "時間 (UTC)",
    "數量 (Shares)",
    "價格 ($)",
    "金額 (USDC)",
    "交易 Hash",
    "說明",
]
col_widths = [46, 16, 8, 22, 22, 14, 10, 12, 68, 80]

# 標題行
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border()

ws.row_dimensions[1].height = 30

# 內容行
COLOR_BY_ACTION = {
    "BUY": "D5E8D4",
    "SELL": "FFE6CC",
    "？": "FFF9C4",
}

for row_idx, f in enumerate(findings, 2):
    addr, name, market, action, date, shares, price, amount, txhash, note = f

    # 判斷行顏色
    if "BUY" in action:
        bg = "D5E8D4"  # 綠
    elif "SELL" in action:
        bg = "FFE6CC"  # 橘
    else:
        bg = "FFF9C4"  # 黃

    vals = [addr, name, market, action, date, shares, price, amount, txhash, note]
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=row_idx, column=col, value=val)
        c.fill = PatternFill("solid", fgColor=bg)
        c.border = border()
        c.alignment = Alignment(vertical="center", wrap_text=(col == len(headers)))

# 空行 + 結論
note_row = len(findings) + 3
ws.cell(note_row, 1, "【調查結論】").font = Font(bold=True, size=12, color="CC0000")

conclusions = [
    "Kuomintang：有完整的 BUY（2025-12-26, $0.84, 100股）→ SELL（2026-02-07, $0.87, 100股）流程。",
    "  但 BUY 未被 Polymarket trades API 收錄（直接合約操作），故只顯示 SELL。淨獲利約 $3。",
    "",
    "ChiangWan-an：代幣來源無 ERC-1155 直接 Transfer 紀錄，疑為 negRisk adapter 內部持倉（買 No → 轉換 Yes）。",
    "  SELL 20股（2026-02-07）和 SELL 80股（2026-04-10）均已確認。",
    "",
    "【共同規律】兩個地址的 BUY 操作都未出現在 Polymarket trades API，",
    "  原因可能是：直接合約交互 / negRisk adapter 操作 / 批次撮合未被單獨記錄。",
]
for i, line in enumerate(conclusions):
    c = ws.cell(note_row + 1 + i, 1, line)
    c.alignment = Alignment(wrap_text=True)
    if line.startswith("  "):
        c.font = Font(color="555555")
    elif line == "":
        pass
    else:
        c.font = Font(bold=True, color="333333")
ws.merge_cells(f"A{note_row}:J{note_row}")

# 欄寬
for i, w in enumerate(col_widths, 1):
    from openpyxl.utils import get_column_letter

    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"

wb.save(EXCEL_PATH)
print(f"已更新: {EXCEL_PATH}")
print(f"已新增「代幣來源調查」工作表")
