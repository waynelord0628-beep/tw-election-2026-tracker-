#!/usr/bin/env python3
"""
monitor.py - 近即時監控 KMT/DPP/TPP 交易
─────────────────────────────────────────
策略：每 POLL_INTERVAL 秒輪詢 Blockscout 一次，
      偵測到新交易立即印出 + 寫入 SQLite + 更新 監控_即時.xlsx

用法：
    python monitor.py           （預設 30 秒輪詢）
    python monitor.py 60        （60 秒輪詢）
"""

import json
import os
import sqlite3
import sys
import time
import asyncio
from datetime import datetime, timezone, timedelta

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── 設定 ──────────────────────────────────────────────────────────────────
POLL_INTERVAL = 10  # 秒

# 自動推算專案根目錄（scripts/ 的上層），相容 Windows / Linux
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(_SCRIPT_DIR)

DB_PATH = os.path.join(PROJECT_ROOT, "monitor.db")
EXCEL_PATH = os.path.join(PROJECT_ROOT, "監控_即時.xlsx")
USDC_CACHE_FILE = os.path.join(PROJECT_ROOT, "usdc_cache.json")
NAME_CACHE_FILE = os.path.join(PROJECT_ROOT, "wallet_names.json")
WEB_DATA_PATH = os.path.join(PROJECT_ROOT, "docs", "data.json")
WEB_FEED_LIMIT = 5000  # data.json 內最多多少筆交易（含查詢用全歷史）
GIT_AUTO_PUSH = True  # 自動 commit + push docs/data.json
GIT_REPO_DIR = PROJECT_ROOT

TZ8 = timezone(timedelta(hours=8))

# ─── 合約地址 ──────────────────────────────────────────────────────────────
CTF = "0x4d97dcd97ec945f40cf65f87097ace5ea0476045"
CTF_ORIGINAL = "0x4D97DCd97eC945f40cF65F87097ACe5EA0476045"
EXCHANGE = "0xc5d563a36ae78145c45a50134d48a1215220f80a"
ADAPTER = "0xd91e80cf2e7be2e162c6513ced06f1dd0da35296"
ZERO = "0x0000000000000000000000000000000000000000"
KNOWN_CONTRACTS = {CTF, EXCHANGE, ADAPTER, ZERO}

# ─── Token 對照表 ──────────────────────────────────────────────────────────
# token_id -> (party, outcome)
TOKENS = {
    "85632914518786177256583369552125280053108667306405854845853340618248288927460": (
        "KMT",
        "Yes",
    ),
    "4696955573632845407532815267539406678302911508204032661527405293140196109387": (
        "KMT",
        "No",
    ),
    "13628189982642424912108657221169198338993179248246381972030640500448717195916": (
        "DPP",
        "Yes",
    ),
    "91004506882941445266754771479824617369805789899332711132070603219216406556613": (
        "DPP",
        "No",
    ),
    "14999500579901383072635205035227864886528710236540822730141548371372688859422": (
        "TPP",
        "Yes",
    ),
    "16222840603445450947154718759167300491302153317593739623696847197718420087623": (
        "TPP",
        "No",
    ),
}
# token_key -> token_id（用於輪詢）
TOKEN_KEY_TO_ID = {f"{p}_{o}": tid for tid, (p, o) in TOKENS.items()}
TOKEN_KEYS_ORDERED = ["KMT_Yes", "KMT_No", "DPP_Yes", "DPP_No", "TPP_Yes", "TPP_No"]

# ─── 知名錢包 ──────────────────────────────────────────────────────────────
KNOWN_WALLETS = {
    "0xd8dd45139269031b16a54717cabad4af6a3980d6": "調查目標-1",
    "0x0d5ee5c536ccd78d179487b3d7e43ae4304d5c24": "ChiangWan-an",
    "0xfde3a53d58320a3db74dbe1092979c401e35719a": "jamieamoy",
    "0x426227d4a9c4ad5a1aae7f2706238f2154b9abaa": "Kuomintang",
    "0xc8ab97a9089a9ff7e6ef0688e6e591a066946418": "ArmageddonRewardsBilly",
    "0xc2358d03312b05b244bde5286dee03bc60ac99f8": "Anon-0xCE",
    "0x0362bb926368b144e0ff98f6828a251e4cb6449e": "varch01",
    "0x0cd7bea497efb9220105858617f0cd660d0a78e0": "SirJason",
    "0x06d248d4f372601d24192284bff919a2c05dfb27": "cheesymm",
    "0xc25120b27e01031b2122f74488dcdb077a78b9c3": "TTbilly",
    "0x864011d381ccdff185896ab92b8173abac1a943c": "ReefTahu",
}

PARTY_COLORS = {
    "KMT": ("002868", "FFFFFF"),
    "DPP": ("1B9431", "FFFFFF"),
    "TPP": ("28C8C8", "000000"),
}

HEADERS = [
    "時間 (UTC+8)",
    "政黨",
    "方向",
    "標的",
    "Shares",
    "價格 ($)",
    "總金額 ($)",
    "交易者",
    "錢包地址",
    "交易 Hash",
    "備註",
]
COL_WIDTHS = [22, 6, 6, 8, 14, 10, 12, 20, 44, 68, 16]
BLUE_DARK = "1F3864"


# ═══════════════════════════════════════════════════════════════════════════
# SQLite
# ═══════════════════════════════════════════════════════════════════════════


def init_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS trades (
            txhash    TEXT,
            token_key TEXT,
            wallet    TEXT,
            direction TEXT,
            shares    REAL,
            price     REAL    DEFAULT 0,
            total     REAL    DEFAULT 0,
            name      TEXT    DEFAULT '',
            timestamp TEXT,
            party     TEXT,
            outcome   TEXT,
            note      TEXT    DEFAULT '',
            PRIMARY KEY (txhash, token_key, wallet, direction)
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ts ON trades(timestamp)")
    conn.commit()
    return conn


# ═══════════════════════════════════════════════════════════════════════════
# USDC 快取
# ═══════════════════════════════════════════════════════════════════════════


def load_usdc_cache() -> dict:
    if os.path.exists(USDC_CACHE_FILE):
        with open(USDC_CACHE_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_usdc_cache(cache: dict):
    with open(USDC_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f)


# ═══════════════════════════════════════════════════════════════════════════
# 用戶名稱快取（自動從 Polymarket 查）
# ═══════════════════════════════════════════════════════════════════════════

_name_cache: dict = {}


def load_name_cache():
    global _name_cache
    if os.path.exists(NAME_CACHE_FILE):
        with open(NAME_CACHE_FILE, encoding="utf-8") as f:
            _name_cache = json.load(f)
    # 把預設知名錢包灌進快取
    for w, n in KNOWN_WALLETS.items():
        _name_cache.setdefault(w.lower(), n)


def save_name_cache():
    with open(NAME_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(_name_cache, f, ensure_ascii=False, indent=2)


def fetch_polymarket_name(wallet: str) -> str:
    """從 Polymarket data-api 查用戶名。失敗回傳空字串。"""
    wallet = wallet.lower()
    try:
        r = requests.get(
            f"https://data-api.polymarket.com/profile?address={wallet}",
            timeout=8,
        )
        if r.status_code == 200:
            d = r.json()
            if isinstance(d, list) and d:
                d = d[0]
            if isinstance(d, dict):
                name = d.get("name") or d.get("pseudonym") or ""
                if name:
                    return name
    except Exception:
        pass

    # 備援：從該 wallet 的最近一筆 trade 撈 name
    try:
        r = requests.get(
            f"https://data-api.polymarket.com/trades?user={wallet}&limit=1",
            timeout=8,
        )
        if r.status_code == 200:
            arr = r.json()
            if arr:
                return arr[0].get("name") or arr[0].get("pseudonym") or ""
    except Exception:
        pass
    return ""


def get_wallet_name(wallet: str) -> str:
    """先查快取，沒有就呼叫 Polymarket API 並寫入快取（含空值，避免重複問）"""
    w = wallet.lower()
    if w in _name_cache:
        return _name_cache[w]
    name = fetch_polymarket_name(w)
    _name_cache[w] = name
    save_name_cache()
    if name:
        print(f"  [新名稱] {w} → {name}")
    return name


# ═══════════════════════════════════════════════════════════════════════════
# 價格查詢（USDC 流向）
# ═══════════════════════════════════════════════════════════════════════════


def fetch_price(txh: str, wallet: str, shares: float, usdc_cache: dict) -> float:
    """從 USDC 流向推算單價，快取避免重複查詢。失敗回傳 0.0"""
    if txh not in usdc_cache:
        try:
            r = requests.get(
                f"https://polygon.blockscout.com/api/v2/transactions/{txh}/token-transfers",
                timeout=15,
            )
            if r.status_code != 200:
                return 0.0
            flows: dict = {}
            for item in r.json().get("items", []):
                sym = (item.get("token", {}).get("symbol") or "").upper()
                if sym not in ("USDC", "USDC.E"):
                    continue
                frm2 = item.get("from", {}).get("hash", "").lower()
                to2 = item.get("to", {}).get("hash", "").lower()
                v = int(item.get("total", {}).get("value", 0)) / 1e6
                if to2 not in KNOWN_CONTRACTS:
                    flows[to2] = flows.get(to2, 0) + v
                if frm2 not in KNOWN_CONTRACTS:
                    flows[frm2] = flows.get(frm2, 0) - v
            usdc_cache[txh] = flows
            save_usdc_cache(usdc_cache)
        except Exception:
            return 0.0

    usdc = usdc_cache[txh].get(wallet, 0)
    if shares > 0 and abs(usdc) > 0:
        p = abs(usdc) / shares
        if 0 < p <= 1.0:
            return round(p, 6)
    return 0.0


# ═══════════════════════════════════════════════════════════════════════════
# 處理單筆 transfer → 寫入 DB
# ═══════════════════════════════════════════════════════════════════════════


def process_transfer(
    conn, txhash, token_key, frm, to_, shares, timestamp_str, usdc_cache
) -> dict | None:
    """
    若為新交易則寫入 DB 並回傳 row dict；
    若已存在或非目標 token 則回傳 None。
    """
    party, outcome = token_key.split("_")

    frm_c = frm in KNOWN_CONTRACTS
    to_c = to_ in KNOWN_CONTRACTS

    if frm_c and to_c:
        return None  # contract-to-contract

    if frm_c:
        wallet, direction = to_, "BUY"
    elif to_c:
        wallet, direction = frm, "SELL"
    else:
        # user-to-user（記 to 為買方）
        wallet, direction = to_, "BUY"

    if shares < 0.000001:
        return None

    # 去重
    cur = conn.execute(
        "SELECT 1 FROM trades WHERE txhash=? AND token_key=? AND wallet=? AND direction=?",
        (txhash, token_key, wallet, direction),
    )
    if cur.fetchone():
        return None

    price = fetch_price(txhash, wallet, shares, usdc_cache)
    total = round(shares * price, 4)
    name = get_wallet_name(wallet)

    conn.execute(
        """INSERT OR IGNORE INTO trades
           (txhash, token_key, wallet, direction, shares, price, total,
            name, timestamp, party, outcome, note)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            txhash,
            token_key,
            wallet,
            direction,
            shares,
            price,
            total,
            name,
            timestamp_str,
            party,
            outcome,
            "",
        ),
    )
    conn.commit()

    return dict(
        txhash=txhash,
        token_key=token_key,
        wallet=wallet,
        direction=direction,
        shares=shares,
        price=price,
        total=total,
        name=name,
        timestamp=timestamp_str,
        party=party,
        outcome=outcome,
        note="",
    )


# ═══════════════════════════════════════════════════════════════════════════
# Excel 重建
# ═══════════════════════════════════════════════════════════════════════════


def ts_display(ts_str: str) -> str:
    try:
        dt = datetime.fromisoformat(ts_str.replace("Z", "+00:00"))
        return dt.astimezone(TZ8).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ts_str


def _write_header_row(ws, row: int = 1):
    for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLUE_DARK)
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = w


def _write_data_row(ws, row_idx: int, db_row: tuple):
    (
        ts_str,
        party,
        direction,
        outcome,
        shares,
        price,
        total,
        name,
        wallet,
        txhash,
        note,
    ) = db_row
    values = [
        ts_display(ts_str),
        party,
        direction,
        outcome,
        shares,
        price,
        total,
        name,
        wallet,
        txhash,
        note,
    ]
    bg, fg = PARTY_COLORS.get(party, ("FFFFFF", "000000"))
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=ci, value=val)
        if ci == 2:
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(color=fg, bold=True)
            c.alignment = Alignment(horizontal="center")
        elif ci == 3:
            c.fill = PatternFill(
                "solid", fgColor="C6EFCE" if val == "BUY" else "FFCCCC"
            )
            c.alignment = Alignment(horizontal="center")
        elif ci == 8:  # 用戶名稱
            if val:
                c.font = Font(color="0066CC", bold=True)
            else:
                c.font = Font(color="999999", italic=True)
        elif ci == 9:  # 錢包地址
            c.font = Font(color="555555", size=9)
        elif ci == 3:
            c.fill = PatternFill(
                "solid", fgColor="C6EFCE" if val == "BUY" else "FFCCCC"
            )
            c.alignment = Alignment(horizontal="center")


def rebuild_excel(conn, *, silent_if_locked: bool = True) -> int | None:
    """完整重建所有分頁。檔案被佔用時 silent_if_locked=True 則靜默跳過。"""
    rows = conn.execute(
        """SELECT timestamp, party, direction, outcome, shares, price, total,
                  name, wallet, txhash, note
           FROM trades ORDER BY timestamp ASC"""
    ).fetchall()

    wb = Workbook()

    # ── AllTrades ──
    ws_all = wb.active
    ws_all.title = "AllTrades"
    ws_all.freeze_panes = "A2"
    _write_header_row(ws_all)
    for ri, row in enumerate(rows, 2):
        _write_data_row(ws_all, ri, row)

    # ── 最新動態（最後 100 筆，倒序）──
    ws_new = wb.create_sheet("最新動態")
    _write_recent_sheet(ws_new, rows[-100:])

    # ── 統計摘要 ──
    ws_stat = wb.create_sheet("統計摘要")
    _build_stat_sheet(ws_stat, rows)

    # ── 各黨分頁 ──
    for party in ("KMT", "DPP", "TPP"):
        ws_p = wb.create_sheet(party)
        ws_p.freeze_panes = "A2"
        _write_header_row(ws_p)
        ri = 2
        for row in rows:
            if row[1] == party:
                _write_data_row(ws_p, ri, row)
                ri += 1

    try:
        wb.save(EXCEL_PATH)
    except PermissionError:
        if silent_if_locked:
            return None
        alt = EXCEL_PATH.replace(".xlsx", f"_{int(time.time())}.xlsx")
        wb.save(alt)
        print(f"  [警告] 主檔被佔用，已另存 {os.path.basename(alt)}")

    return len(rows)


def update_recent_excel(conn) -> int | None:
    """
    輕量更新：只寫一個「最新動態.xlsx」獨立小檔，
    平常每次有新交易都呼叫這個（毫秒級）。
    """
    rows = conn.execute(
        """SELECT timestamp, party, direction, outcome, shares, price, total,
                  name, wallet, txhash, note
           FROM trades ORDER BY timestamp DESC LIMIT 100"""
    ).fetchall()

    recent_path = EXCEL_PATH.replace(".xlsx", "_最新動態.xlsx")
    wb = Workbook()
    ws = wb.active
    if ws is None:
        return None
    ws.title = "最新動態"
    _write_recent_sheet(ws, [r for r in reversed(rows)])

    try:
        wb.save(recent_path)
    except PermissionError:
        return None
    return len(rows)


def _write_recent_sheet(ws, rows_chrono):
    """rows_chrono 為時間順序（舊→新）；本分頁顯示為新→舊"""
    ws.freeze_panes = "A3"
    now_str = datetime.now(TZ8).strftime("%Y-%m-%d %H:%M:%S")
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = f"最新動態（最後 {len(rows_chrono)} 筆）　最後更新：{now_str}"
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=BLUE_DARK)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22
    _write_header_row(ws, row=2)
    for ri, row in enumerate(reversed(rows_chrono), 3):
        _write_data_row(ws, ri, row)


def _build_stat_sheet(ws, rows):
    from collections import defaultdict

    ws["A1"] = "統計項目"
    ws["A1"].font = Font(bold=True)
    ws["B1"] = "數值"
    ws["B1"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    party_buy = defaultdict(float)
    party_sell = defaultdict(float)
    party_vol = defaultdict(float)

    for row in rows:
        _, party, direction, _, shares, _, total, *_ = row
        party_vol[party] += total
        if direction == "BUY":
            party_buy[party] += shares
        else:
            party_sell[party] += shares

    stats = [
        ("總交易筆數", len(rows)),
        ("", ""),
        ("KMT 交易筆數", sum(1 for r in rows if r[1] == "KMT")),
        ("KMT BUY 總 Shares", round(party_buy["KMT"], 2)),
        ("KMT SELL 總 Shares", round(party_sell["KMT"], 2)),
        ("KMT 總交易量 (USD)", round(party_vol["KMT"], 2)),
        ("", ""),
        ("DPP 交易筆數", sum(1 for r in rows if r[1] == "DPP")),
        ("DPP BUY 總 Shares", round(party_buy["DPP"], 2)),
        ("DPP SELL 總 Shares", round(party_sell["DPP"], 2)),
        ("DPP 總交易量 (USD)", round(party_vol["DPP"], 2)),
        ("", ""),
        ("TPP 交易筆數", sum(1 for r in rows if r[1] == "TPP")),
        ("TPP BUY 總 Shares", round(party_buy["TPP"], 2)),
        ("TPP SELL 總 Shares", round(party_sell["TPP"], 2)),
        ("TPP 總交易量 (USD)", round(party_vol["TPP"], 2)),
        ("", ""),
        ("最後更新", datetime.now(TZ8).strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for ri, (k, v) in enumerate(stats, 2):
        ws.cell(row=ri, column=1, value=k)
        ws.cell(row=ri, column=2, value=v)


# ═══════════════════════════════════════════════════════════════════════════
# 終端機輸出
# ═══════════════════════════════════════════════════════════════════════════


def log_new_trade(trade: dict):
    ts = ts_display(trade["timestamp"])
    raw_name = trade["name"]
    name = f"\033[1;34m{raw_name}\033[0m" if raw_name else "\033[2;37m(未命名)\033[0m"
    p = f"${trade['price']:.4f}" if trade["price"] else "未知價"
    dir_icon = "▲ BUY " if trade["direction"] == "BUY" else "▼ SELL"
    total = f"${trade['total']:>8.2f}" if trade["total"] else "    -    "
    print(
        f"  [{ts}] {trade['party']}-{trade['outcome']:3s} | "
        f"{dir_icon} {trade['shares']:>10.2f} @ {p:<10} = {total}"
    )
    print(f"      用戶：{name}")
    print(f"      地址：{trade['wallet']}")
    print(f"      Hash：{trade['txhash']}")


# ═══════════════════════════════════════════════════════════════════════════
# 公開網站資料匯出（web/data.json）
# ═══════════════════════════════════════════════════════════════════════════


def export_web_data(conn) -> int | None:
    """
    把 SQLite 資料匯出成 web/data.json，給 GitHub Pages 網頁使用。
    包含：updated_at、總筆數、各黨統計、最新 WEB_FEED_LIMIT 筆交易。
    """
    from collections import defaultdict

    rows = conn.execute(
        """SELECT timestamp, party, direction, outcome, shares, price, total,
                  name, wallet, txhash, note
           FROM trades ORDER BY timestamp DESC"""
    ).fetchall()

    # 統計
    stats: dict = {
        p: {"count": 0, "buy_shares": 0.0, "sell_shares": 0.0, "volume_usd": 0.0}
        for p in ("KMT", "DPP", "TPP")
    }
    for r in rows:
        _, party, direction, _, shares, _, total, *_ = r
        if party not in stats:
            continue
        stats[party]["count"] += 1
        stats[party]["volume_usd"] += total or 0
        if direction == "BUY":
            stats[party]["buy_shares"] += shares or 0
        else:
            stats[party]["sell_shares"] += shares or 0
    for p in stats:
        stats[p]["buy_shares"] = round(stats[p]["buy_shares"], 2)
        stats[p]["sell_shares"] = round(stats[p]["sell_shares"], 2)
        stats[p]["volume_usd"] = round(stats[p]["volume_usd"], 2)

    # 最新 N 筆（已是 DESC）
    feed = []
    for r in rows[:WEB_FEED_LIMIT]:
        ts, party, direction, outcome, shares, price, total, name, wallet, txh, note = r
        feed.append(
            {
                "timestamp": ts,
                "party": party,
                "direction": direction,
                "outcome": outcome,
                "shares": round(shares or 0, 4),
                "price": round(price or 0, 6),
                "total": round(total or 0, 4),
                "name": name or "",
                "wallet": wallet,
                "txhash": txh,
                "note": note or "",
            }
        )

    payload = {
        "updated_at": datetime.now(TZ8).strftime("%Y-%m-%d %H:%M:%S"),
        "total_count": len(rows),
        "stats": stats,
        "trades": feed,
    }

    try:
        os.makedirs(os.path.dirname(WEB_DATA_PATH), exist_ok=True)
        # 原子寫入：先寫 tmp 再 rename，避免網頁讀到一半的檔
        tmp = WEB_DATA_PATH + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
        os.replace(tmp, WEB_DATA_PATH)
    except Exception as e:
        print(f"  [web 匯出失敗] {e}")
        return None
    return len(rows)


# ═══════════════════════════════════════════════════════════════════════════
# Git auto-push（節流：每 GIT_PUSH_INTERVAL 秒最多 push 一次）
# ═══════════════════════════════════════════════════════════════════════════

import subprocess

GIT_PUSH_INTERVAL = 120  # 秒；最少間隔避免 commit 爆量
_last_git_push = 0.0


def git_push_data(new_count: int = 0) -> bool:
    """commit & push docs/data.json。失敗不影響主流程。"""
    global _last_git_push
    if not GIT_AUTO_PUSH:
        return False
    now = time.time()
    if now - _last_git_push < GIT_PUSH_INTERVAL:
        return False
    try:
        # 確認是 git repo
        chk = subprocess.run(
            ["git", "rev-parse", "--is-inside-work-tree"],
            cwd=GIT_REPO_DIR,
            capture_output=True,
            text=True,
            timeout=10,
        )
        if chk.returncode != 0:
            return False

        subprocess.run(
            ["git", "add", "docs/data.json"],
            cwd=GIT_REPO_DIR,
            capture_output=True,
            timeout=15,
        )
        # 若無變動 commit 會回傳非 0，不視為錯誤
        msg = f"data: +{new_count} trades @ {datetime.now(TZ8).strftime('%H:%M:%S')}"
        cm = subprocess.run(
            ["git", "commit", "-m", msg],
            cwd=GIT_REPO_DIR,
            capture_output=True,
            text=True,
            timeout=20,
        )
        if "nothing to commit" in (cm.stdout + cm.stderr):
            _last_git_push = now
            return False

        ps = subprocess.run(
            ["git", "push"],
            cwd=GIT_REPO_DIR,
            capture_output=True,
            text=True,
            timeout=60,
        )
        _last_git_push = now
        if ps.returncode == 0:
            print(f"  [git] 已推送 → GitHub Pages（{msg}）")
            return True
        else:
            print(f"  [git push 失敗] {ps.stderr.strip()[:200]}")
            return False
    except Exception as e:
        print(f"  [git 錯誤] {e}")
        return False


# ═══════════════════════════════════════════════════════════════════════════
# 初始化：從 Blockscout 拉取所有歷史資料
# ═══════════════════════════════════════════════════════════════════════════


def import_history(conn, usdc_cache: dict):
    count = conn.execute("SELECT COUNT(*) FROM trades").fetchone()[0]
    if count > 0:
        print(f"[DB] 已有 {count} 筆資料，跳過歷史匯入")
        return

    print("[初始化] DB 為空，正在從 Blockscout 匯入全部歷史資料...")
    total_new = 0
    for token_key in TOKEN_KEYS_ORDERED:
        tid = TOKEN_KEY_TO_ID[token_key]
        url = (
            f"https://polygon.blockscout.com/api/v2/tokens/"
            f"{CTF_ORIGINAL}/instances/{tid}/transfers"
        )
        page_params = None
        batch = 0
        print(f"  {token_key}...", end="", flush=True)
        while True:
            try:
                r = requests.get(url, params=page_params, timeout=30)
                if r.status_code != 200:
                    break
                d = r.json()
            except Exception as e:
                print(f" [ERR:{e}]", end="")
                break
            for item in d.get("items", []):
                frm = item.get("from", {}).get("hash", "").lower()
                to_ = item.get("to", {}).get("hash", "").lower()
                txh = item.get("transaction_hash", "").lower()
                ts = item.get("timestamp", "")
                val = int(item.get("total", {}).get("value", 0)) / 1e6
                res = process_transfer(
                    conn, txh, token_key, frm, to_, val, ts, usdc_cache
                )
                if res:
                    batch += 1
            nxt = d.get("next_page_params")
            if not nxt:
                break
            page_params = nxt
            time.sleep(0.3)
        print(f" +{batch}")
        total_new += batch

    print(f"[初始化] 完成，共匯入 {total_new} 筆")


# ═══════════════════════════════════════════════════════════════════════════
# 主輪詢迴圈
# ═══════════════════════════════════════════════════════════════════════════

# 每個 token_key 記錄最新一頁的 txhash set，用於快速去重
_last_seen: dict[str, set] = {k: set() for k in TOKEN_KEYS_ORDERED}


def poll_once(conn, usdc_cache: dict) -> int:
    """
    對所有 6 個 token 輪詢；自動翻頁直到遇到已知 hash 為止，避免爆量漏單。
    """
    new_total = 0
    MAX_PAGES = 20  # 安全上限，避免無限翻頁

    for token_key in TOKEN_KEYS_ORDERED:
        tid = TOKEN_KEY_TO_ID[token_key]
        url = (
            f"https://polygon.blockscout.com/api/v2/tokens/"
            f"{CTF_ORIGINAL}/instances/{tid}/transfers"
        )

        page_params = None
        first_page_hashes: set = set()
        seen_known = False
        pages_fetched = 0

        while pages_fetched < MAX_PAGES:
            try:
                r = requests.get(url, params=page_params, timeout=20)
                if r.status_code != 200:
                    break
                d = r.json()
            except Exception as e:
                print(f"  [輪詢錯誤 {token_key}] {e}")
                break

            items = d.get("items", [])
            if not items:
                break

            for item in items:
                frm = item.get("from", {}).get("hash", "").lower()
                to_ = item.get("to", {}).get("hash", "").lower()
                txh = item.get("transaction_hash", "").lower()
                ts = item.get("timestamp", "")
                val = int(item.get("total", {}).get("value", 0)) / 1e6

                if pages_fetched == 0:
                    first_page_hashes.add(txh)

                # 若這個 hash 上輪見過 → 證明追上了，停止翻頁
                if txh in _last_seen[token_key]:
                    seen_known = True
                    continue

                res = process_transfer(
                    conn, txh, token_key, frm, to_, val, ts, usdc_cache
                )
                if res:
                    log_new_trade(res)
                    new_total += 1

            pages_fetched += 1

            # 已遇到上輪看過的 hash → 不必再往下翻
            if seen_known:
                break
            # 第一次啟動（_last_seen 為空）只抓第一頁，避免重抓全歷史
            if not _last_seen[token_key]:
                break

            nxt = d.get("next_page_params")
            if not nxt:
                break
            page_params = nxt
            time.sleep(0.2)

        # 更新 _last_seen 為「本輪第一頁」的 hash 集合
        if first_page_hashes:
            _last_seen[token_key] = first_page_hashes

        if pages_fetched > 1:
            print(f"  [翻頁] {token_key} 抓了 {pages_fetched} 頁追上進度")

    return new_total


# ═══════════════════════════════════════════════════════════════════════════
# 主程式
# ═══════════════════════════════════════════════════════════════════════════


def main(poll_interval: int = POLL_INTERVAL):
    print("=" * 65)
    print("  KMT / DPP / TPP 交易監控")
    print(f"  輪詢間隔：{poll_interval} 秒　　輸出：{os.path.basename(EXCEL_PATH)}")
    print("  Ctrl+C 停止")
    print("=" * 65)

    conn = init_db()
    usdc_cache = load_usdc_cache()
    load_name_cache()
    print(f"[名稱快取] 已載入 {len(_name_cache)} 個錢包名稱")

    # 首次執行：拉歷史、建 Excel
    import_history(conn, usdc_cache)
    total = rebuild_excel(conn, silent_if_locked=False)
    print(f"[Excel] 初始建立完成：{total} 筆 → {EXCEL_PATH}")
    web_n = export_web_data(conn)
    if web_n is not None:
        print(f"[Web ] 初始 data.json 已產生：{web_n} 筆 → {WEB_DATA_PATH}")
    # 啟動時強制 push 一次，更新網站「最後更新」時間戳，並驗證 git auth
    global _last_git_push
    _last_git_push = 0  # 解除節流
    if git_push_data(0):
        print(f"[Git ] 初始 push 成功")
    else:
        print(f"[Git ] 初始 push 失敗或被節流（檢查 git auth）")
    print("-" * 65)
    print(f"[監控中] 每 {poll_interval} 秒輪詢一次 Blockscout...")
    print()

    REBUILD_EVERY = 600  # 秒；完整重建主檔的間隔
    last_rebuild = time.time()
    cycle = 0

    while True:
        try:
            cycle += 1
            new_count = poll_once(conn, usdc_cache)

            if new_count > 0:
                # 平常：只更新「最新動態」獨立小檔（毫秒級、永不卡）
                update_recent_excel(conn)
                export_web_data(conn)  # 同步給網頁
                git_push_data(new_count)  # 節流推送至 GitHub Pages
                total = conn.execute("SELECT COUNT(*) FROM trades").fetchone()[0]
                print(f"  └─ [更新] 累計 {total} 筆（本輪 +{new_count}）")
            else:
                if cycle % 10 == 0:
                    now = datetime.now(TZ8).strftime("%H:%M:%S")
                    total = conn.execute("SELECT COUNT(*) FROM trades").fetchone()[0]
                    print(f"[{now}] 輪詢 #{cycle}，無新交易（累計 {total} 筆）")

            # 每 10 分鐘完整重建主檔（含全分頁）；被佔用就靜默跳過
            if time.time() - last_rebuild >= REBUILD_EVERY:
                result = rebuild_excel(conn, silent_if_locked=True)
                if result is not None:
                    print(f"  [完整重建] 主檔已更新（{result} 筆）")
                else:
                    print(f"  [完整重建] 主檔被開啟中，下次再試")
                last_rebuild = time.time()

            time.sleep(poll_interval)

        except KeyboardInterrupt:
            print("\n[停止] 監控已結束")
            break
        except Exception as e:
            print(f"[錯誤] {e}，10 秒後繼續...")
            time.sleep(10)


if __name__ == "__main__":
    interval = int(sys.argv[1]) if len(sys.argv) > 1 else POLL_INTERVAL
    main(interval)
