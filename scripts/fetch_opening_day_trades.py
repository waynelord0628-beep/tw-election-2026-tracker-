import requests
import time
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TZ8 = timezone(timedelta(hours=8))

# Market conditionIds
MARKETS = {
    "KMT": "0xc0f076bc4d90a44df34a729277e9d1f294f0cb60d2c3b1b3800908b1e15b923b",
    "DPP": "0xea3b1d0099085f43a3098b3e1fbcbe62284ce1bda99384b3d46b82ff202ac016",
    "TPP": "0x68fc8d466ddc10a1ae37b52642f36b93e413cf98ba5fe3947c242a9a727b2e94",
}

# Dec 4–6, 2025 UTC range
DAY_START = 1764806400  # 2025-12-04 00:00:00 UTC
DAY_END = 1765151999  # 2025-12-06 23:59:59 UTC


def fetch_all_trades(condition_id, party):
    """Fetch all trades for a market, return only opening day trades."""
    trades = []
    limit = 100
    offset = 0
    print(f"Fetching {party} trades...")
    while True:
        url = f"https://data-api.polymarket.com/trades?market={condition_id}&limit={limit}&offset={offset}"
        resp = requests.get(url, timeout=30)
        if resp.status_code != 200:
            print(f"  Error {resp.status_code} at offset {offset}")
            break
        batch = resp.json()
        if not batch:
            break
        for t in batch:
            ts = t.get("timestamp", 0)
            if DAY_START <= ts <= DAY_END:
                trades.append(t)
            # Since sorted descending, once we go below DAY_START we can stop
            elif ts < DAY_START:
                print(f"  Reached before opening day at offset {offset}")
                return trades
        offset += limit
        time.sleep(0.2)
    print(f"  Found {len(trades)} trades (12/4–12/6) for {party}")
    return trades


def ts_to_dt(ts):
    return datetime.fromtimestamp(ts, tz=TZ8).strftime("%Y-%m-%d %H:%M:%S")


def build_excel(all_data):
    wb = Workbook()

    party_colors = {
        "KMT": "002868",  # KMT blue
        "DPP": "1B9431",  # DPP green
        "TPP": "28C8C8",  # TPP teal
    }
    header_font_color = {
        "KMT": "FFFFFF",
        "DPP": "FFFFFF",
        "TPP": "000000",
    }

    headers = [
        "時間 (UTC+8)",
        "政黨",
        "方向",
        "標的 (Yes/No)",
        "數量 (Shares)",
        "價格 ($)",
        "總金額 ($)",
        "交易者名稱",
        "錢包地址",
        "交易 Hash",
    ]
    col_widths = [22, 6, 6, 14, 16, 10, 12, 20, 44, 68]

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Summary sheet
    ws_all = wb.active
    ws_all.title = "全部交易"

    # Per-party sheets
    sheets = {"全部交易": ws_all}
    for party in ["KMT", "DPP", "TPP"]:
        ws = wb.create_sheet(title=party)
        sheets[party] = ws

    # Write to each sheet
    for sheet_name, ws in sheets.items():
        # Header row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(fill_type="solid", fgColor="1F3864")
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[
                col_idx - 1
            ]
        ws.row_dimensions[1].height = 30

        row_idx = 2

        # 全部交易分頁：跨市場統一按時間排序
        if sheet_name == "全部交易":
            flat = []
            for p, ts_list in all_data.items():
                for trade in ts_list:
                    flat.append((p, trade))
            flat.sort(key=lambda x: int(x[1].get("timestamp", 0)))
            iter_trades = flat
        else:
            iter_trades = [(sheet_name, t) for t in all_data[sheet_name]]

        for i, (party, t) in enumerate(iter_trades):
            color = party_colors[party]
            ts = t.get("timestamp", 0)
            side = t.get("side", "")
            outcome = t.get("outcome", "")
            size = round(t.get("size", 0), 6)
            price = round(t.get("price", 0), 4)
            total = round(size * price, 4)
            name = t.get("name", "") or t.get("pseudonym", "")
            wallet = t.get("proxyWallet", "")
            tx_hash = t.get("transactionHash", "")

            row_data = [
                ts_to_dt(ts),
                party,
                side,
                outcome,
                size,
                price,
                total,
                name,
                wallet,
                tx_hash,
            ]

            # Alternate row fill
            bg = "EBF5FB" if i % 2 == 0 else "FFFFFF"

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                cell.border = border
                # Highlight party column
                if col_idx == 2:
                    cell.fill = PatternFill(fill_type="solid", fgColor=color)
                    cell.font = Font(bold=True, color=header_font_color[party], size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
                    cell.font = Font(size=10)
                # Color BUY/SELL
                if col_idx == 3:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if side == "BUY":
                        cell.font = Font(color="006400", bold=True, size=10)
                    else:
                        cell.font = Font(color="8B0000", bold=True, size=10)
                # Color Yes/No
                if col_idx == 4:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            row_idx += 1

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Summary stats sheet
    ws_stat = wb.create_sheet(title="統計摘要")
    stat_headers = [
        "政黨",
        "總交易筆數",
        "BUY 筆數",
        "SELL 筆數",
        "Yes 總量",
        "No 總量",
        "總成交金額 ($)",
        "平均價格 ($)",
    ]
    for col_idx, h in enumerate(stat_headers, 1):
        cell = ws_stat.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(fill_type="solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_stat.column_dimensions[get_column_letter(col_idx)].width = 18

    for row_idx, (party, trades) in enumerate(all_data.items(), 2):
        buys = [t for t in trades if t.get("side") == "BUY"]
        sells = [t for t in trades if t.get("side") == "SELL"]
        yes_vol = sum(t.get("size", 0) for t in trades if t.get("outcome") == "Yes")
        no_vol = sum(t.get("size", 0) for t in trades if t.get("outcome") == "No")
        total_val = sum(t.get("size", 0) * t.get("price", 0) for t in trades)
        avg_price = (
            (sum(t.get("price", 0) for t in trades) / len(trades)) if trades else 0
        )

        row_data = [
            party,
            len(trades),
            len(buys),
            len(sells),
            round(yes_vol, 2),
            round(no_vol, 2),
            round(total_val, 2),
            round(avg_price, 4),
        ]
        color = party_colors[party]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_stat.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if col_idx == 1:
                cell.fill = PatternFill(fill_type="solid", fgColor=color)
                cell.font = Font(bold=True, color=header_font_color[party], size=11)
            else:
                cell.fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
                cell.font = Font(size=11)

    output_path = "E:\\polymarket選舉賭博\\2026台灣地方選舉_開市三日交易.xlsx"
    wb.save(output_path)
    print(f"\nExcel 儲存至: {output_path}")
    return output_path


# Main
all_data = {}
for party, cid in MARKETS.items():
    trades = fetch_all_trades(cid, party)
    # Sort oldest first
    trades.sort(key=lambda x: x["timestamp"])
    all_data[party] = trades
    total = sum(len(v) for v in all_data.values())

print(f"\n總計開市三日（12/4–12/6）交易筆數: {sum(len(v) for v in all_data.values())}")
for p, t in all_data.items():
    print(f"  {p}: {len(t)} 筆")

build_excel(all_data)
