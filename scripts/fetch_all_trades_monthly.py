"""
fetch_all_trades_monthly.py
抓取台灣地方選舉三個市場（KMT/DPP/TPP）所有交易，
產生 Excel：
  - 總表（所有交易，按時間升序）
  - 月份工作表（2025-12、2026-01 … 至今）
  - 統計摘要
"""

import requests
import time
from datetime import datetime, timezone, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TZ8 = timezone(timedelta(hours=8))

MARKETS = {
    "KMT": "0xc0f076bc4d90a44df34a729277e9d1f294f0cb60d2c3b1b3800908b1e15b923b",
    "DPP": "0xea3b1d0099085f43a3098b3e1fbcbe62284ce1bda99384b3d46b82ff202ac016",
    "TPP": "0x68fc8d466ddc10a1ae37b52642f36b93e413cf98ba5fe3947c242a9a727b2e94",
}

PARTY_COLORS = {
    "KMT": ("002868", "FFFFFF"),  # bg, fg
    "DPP": ("1B9431", "FFFFFF"),
    "TPP": ("28C8C8", "000000"),
}

HEADERS = [
    "時間 (UTC+8)",
    "政黨",
    "方向",
    "標的 (Yes/No)",
    "數量 (Shares)",
    "價格 ($)",
    "總金額 ($)",
    "交易者名稱",
    "錢包地址",
    "交易 Hash",
]
COL_WIDTHS = [22, 6, 6, 14, 16, 10, 12, 20, 44, 68]


# ── 抓取 ────────────────────────────────────────────────────────────────────


def fetch_all_trades(condition_id, party):
    trades = []
    limit = 500
    offset = 0
    print(f"  Fetching {party}...", end="", flush=True)
    while True:
        url = (
            f"https://data-api.polymarket.com/trades"
            f"?market={condition_id}&limit={limit}&offset={offset}"
        )
        resp = requests.get(url, timeout=30)
        if resp.status_code != 200:
            print(f" [Error {resp.status_code} at offset {offset}]")
            break
        batch = resp.json()
        if not batch:
            break
        trades.extend(batch)
        offset += limit
        time.sleep(0.25)
        if len(batch) < limit:
            break
    print(f" {len(trades)} 筆")
    return trades


# ── 工具 ────────────────────────────────────────────────────────────────────


def ts_to_dt8(ts):
    return datetime.fromtimestamp(ts, tz=TZ8).strftime("%Y-%m-%d %H:%M:%S")


def month_key(ts):
    """Returns e.g. '2025-12' in UTC+8."""
    return datetime.fromtimestamp(ts, tz=TZ8).strftime("%Y-%m")


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


# ── Excel 樣式輔助 ───────────────────────────────────────────────────────────


def write_header(ws, border):
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(fill_type="solid", fgColor="1F3864")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


def write_trade_rows(ws, trade_tuples, border):
    """
    trade_tuples: list of (party, trade_dict)
    Returns row count written.
    """
    for i, (party, t) in enumerate(trade_tuples):
        row_idx = i + 2
        bg_color, _ = PARTY_COLORS[party]
        ts = t.get("timestamp", 0)
        side = t.get("side", "")
        outcome = t.get("outcome", "")
        size = round(float(t.get("size", 0)), 6)
        price = round(float(t.get("price", 0)), 4)
        total = round(size * price, 4)
        name = t.get("name", "") or t.get("pseudonym", "") or ""
        wallet = t.get("proxyWallet", "")
        tx_hash = t.get("transactionHash", "")

        row_data = [
            ts_to_dt8(ts),
            party,
            side,
            outcome,
            size,
            price,
            total,
            name,
            wallet,
            tx_hash,
        ]

        alt_bg = "EBF5FB" if i % 2 == 0 else "FFFFFF"

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            if col_idx == 2:  # 政黨欄
                cell.fill = PatternFill(fill_type="solid", fgColor=bg_color)
                _, fg = PARTY_COLORS[party]
                cell.font = Font(bold=True, color=fg, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 3:  # BUY/SELL
                color = "006400" if side == "BUY" else "8B0000"
                cell.font = Font(color=color, bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(fill_type="solid", fgColor=alt_bg)
            elif col_idx == 4:  # Yes/No
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(fill_type="solid", fgColor=alt_bg)
            else:
                cell.font = Font(size=10)
                cell.fill = PatternFill(fill_type="solid", fgColor=alt_bg)

    return len(trade_tuples)


# ── 統計摘要工作表 ────────────────────────────────────────────────────────────


def write_summary_sheet(wb, all_trades_by_party, monthly_keys, border):
    ws = wb.create_sheet(title="統計摘要")
    stat_headers = [
        "政黨",
        "總筆數",
        "BUY 筆數",
        "SELL 筆數",
        "Yes 總量",
        "No 總量",
        "總成交金額 ($)",
        "平均價格 ($)",
    ]
    col_widths_s = [8, 10, 10, 10, 14, 14, 18, 14]
    for col_idx, h in enumerate(stat_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(fill_type="solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths_s[
            col_idx - 1
        ]
    ws.row_dimensions[1].height = 28

    for row_idx, (party, trades) in enumerate(all_trades_by_party.items(), 2):
        buys = [t for t in trades if t.get("side") == "BUY"]
        sells = [t for t in trades if t.get("side") == "SELL"]
        yes_vol = sum(
            float(t.get("size", 0)) for t in trades if t.get("outcome") == "Yes"
        )
        no_vol = sum(
            float(t.get("size", 0)) for t in trades if t.get("outcome") == "No"
        )
        total_val = sum(
            float(t.get("size", 0)) * float(t.get("price", 0)) for t in trades
        )
        avg_price = (
            (sum(float(t.get("price", 0)) for t in trades) / len(trades))
            if trades
            else 0
        )

        row_data = [
            party,
            len(trades),
            len(buys),
            len(sells),
            round(yes_vol, 2),
            round(no_vol, 2),
            round(total_val, 2),
            round(avg_price, 4),
        ]
        bg_color, fg_color = PARTY_COLORS[party]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if col_idx == 1:
                cell.fill = PatternFill(fill_type="solid", fgColor=bg_color)
                cell.font = Font(bold=True, color=fg_color, size=11)
            else:
                cell.fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
                cell.font = Font(size=11)

    # 月份摘要
    ws.cell(row=6, column=1, value="月份").font = Font(bold=True, size=11)
    ws.cell(row=6, column=2, value="總筆數").font = Font(bold=True, size=11)
    ws.cell(row=6, column=3, value="KMT").font = Font(bold=True, size=11)
    ws.cell(row=6, column=4, value="DPP").font = Font(bold=True, size=11)
    ws.cell(row=6, column=5, value="TPP").font = Font(bold=True, size=11)
    for col in range(1, 6):
        cell = ws.cell(row=6, column=col)
        cell.fill = PatternFill(fill_type="solid", fgColor="2E4057")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Collect monthly counts
    monthly_counts = defaultdict(lambda: defaultdict(int))
    for party, trades in all_trades_by_party.items():
        for t in trades:
            mk = month_key(t.get("timestamp", 0))
            monthly_counts[mk][party] += 1

    for r_off, mk in enumerate(monthly_keys, 7):
        row_data_m = [
            mk,
            sum(monthly_counts[mk].values()),
            monthly_counts[mk].get("KMT", 0),
            monthly_counts[mk].get("DPP", 0),
            monthly_counts[mk].get("TPP", 0),
        ]
        for col_idx, val in enumerate(row_data_m, 1):
            cell = ws.cell(row=r_off, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            cell.fill = PatternFill(fill_type="solid", fgColor="F9F9F9")
            cell.font = Font(size=10)


# ── 主程式 ───────────────────────────────────────────────────────────────────

print("=== 抓取三市場所有交易 ===")
all_trades_by_party = {}
for party, cid in MARKETS.items():
    trades = fetch_all_trades(cid, party)
    trades.sort(key=lambda x: x.get("timestamp", 0))
    all_trades_by_party[party] = trades

# 合併全部並按時間排序
all_flat = []
for party, trades in all_trades_by_party.items():
    for t in trades:
        all_flat.append((party, t))
all_flat.sort(key=lambda x: x[1].get("timestamp", 0))

total_count = len(all_flat)
print(f"\n總計 {total_count} 筆交易")
for party, trades in all_trades_by_party.items():
    print(f"  {party}: {len(trades)} 筆")

# 按 UTC+8 月份分組
monthly_groups = defaultdict(list)
for party, t in all_flat:
    mk = month_key(t.get("timestamp", 0))
    monthly_groups[mk].append((party, t))

monthly_keys = sorted(monthly_groups.keys())
print(f"\n月份範圍: {monthly_keys[0]} ～ {monthly_keys[-1]}")
for mk in monthly_keys:
    print(f"  {mk}: {len(monthly_groups[mk])} 筆")

# ── 建 Excel ────────────────────────────────────────────────────────────────
print("\n建立 Excel...")
wb = Workbook()
border = thin_border()

# 1. 總表
ws_all = wb.active
ws_all.title = "總表"
write_header(ws_all, border)
write_trade_rows(ws_all, all_flat, border)
print(f"  總表: {len(all_flat)} 筆 ✓")

# 2. 月份工作表
for mk in monthly_keys:
    ws_m = wb.create_sheet(title=mk)
    write_header(ws_m, border)
    rows = monthly_groups[mk]
    write_trade_rows(ws_m, rows, border)
    print(f"  {mk}: {len(rows)} 筆 ✓")

# 3. 統計摘要（放在最後）
write_summary_sheet(wb, all_trades_by_party, monthly_keys, border)
print("  統計摘要 ✓")

output_path = "E:\\polymarket選舉賭博\\2026台灣地方選舉_全部交易.xlsx"
wb.save(output_path)
print(f"\nExcel 儲存至: {output_path}")
