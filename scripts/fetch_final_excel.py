"""
fetch_final_excel.py
最終版：僅用 Polymarket API (964筆)，確認無鏈上補錄需求後直接輸出 Excel
─────────────────────────────────────────────────────────────────
調查結論：
  - Polymarket API 共 964 筆唯一交易（真實方向性交易）
  - Blockscout 上額外的 87 筆均為 split/merge/內部操作（price≈1.0），非方向性交易
  - 無需鏈上補錄，964 筆即為完整資料集

Excel結構：
  總表 / 2025-12 / 2026-01 / ... / 2026-04 / 統計摘要
"""

import requests, time
from datetime import datetime, timezone, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TZ8 = timezone(timedelta(hours=8))

MARKETS = {
    "KMT": "0xc0f076bc4d90a44df34a729277e9d1f294f0cb60d2c3b1b3800908b1e15b923b",
    "DPP": "0xea3b1d0099085f43a3098b3e1fbcbe62284ce1bda99384b3d46b82ff202ac016",
    "TPP": "0x68fc8d466ddc10a1ae37b52642f36b93e413cf98ba5fe3947c242a9a727b2e94",
}

PARTY_COLORS = {
    "KMT": ("002868", "FFFFFF"),
    "DPP": ("1B9431", "FFFFFF"),
    "TPP": ("28C8C8", "000000"),
}

HEADERS = [
    "時間 (UTC+8)",
    "政黨",
    "方向",
    "標的 (Yes/No)",
    "數量 (Shares)",
    "價格 ($)",
    "總金額 ($)",
    "交易者名稱",
    "錢包地址",
    "交易 Hash",
]
COL_WIDTHS = [22, 6, 6, 14, 16, 10, 12, 20, 44, 68]


def fetch_poly_trades(condition_id, party):
    trades = []
    offset = 0
    print(f"  {party}...", end="", flush=True)
    while True:
        r = requests.get(
            f"https://data-api.polymarket.com/trades?market={condition_id}&limit=500&offset={offset}",
            timeout=30,
        )
        batch = r.json()
        if not batch:
            break
        trades.extend(batch)
        offset += 500
        time.sleep(0.25)
        if len(batch) < 500:
            break
    print(f" {len(trades)}筆")
    return trades


def ts_to_dt8(ts):
    return datetime.fromtimestamp(ts, tz=TZ8).strftime("%Y-%m-%d %H:%M:%S")


def month_key(ts):
    return datetime.fromtimestamp(ts, tz=TZ8).strftime("%Y-%m")


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def write_header(ws, border):
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(fill_type="solid", fgColor="1F3864")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


def write_rows(ws, trades, border):
    for i, t in enumerate(trades):
        row_idx = i + 2
        party = t["party"]
        side = t["side"]
        bg_color, fg_color = PARTY_COLORS.get(party, ("888888", "FFFFFF"))
        alt_bg = "EBF5FB" if i % 2 == 0 else "FFFFFF"

        row_data = [
            ts_to_dt8(t["timestamp"]) if t["timestamp"] else "",
            party,
            side,
            t["outcome"],
            t["size"],
            t["price"],
            t["total"],
            t["name"],
            t["proxyWallet"],
            t["transactionHash"],
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            if col_idx == 2:
                cell.fill = PatternFill(fill_type="solid", fgColor=bg_color)
                cell.font = Font(bold=True, color=fg_color, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 3:
                color = "006400" if side == "BUY" else "8B0000"
                cell.font = Font(color=color, bold=True, size=10)
                cell.fill = PatternFill(fill_type="solid", fgColor=alt_bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 4:
                cell.font = Font(size=10)
                cell.fill = PatternFill(fill_type="solid", fgColor=alt_bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = Font(size=10)
                cell.fill = PatternFill(fill_type="solid", fgColor=alt_bg)


def write_summary_sheet(wb, all_trades, monthly_keys, monthly, border):
    ws = wb.create_sheet(title="統計摘要")
    stat_headers = [
        "政黨",
        "總筆數",
        "BUY",
        "SELL",
        "Yes總量(shares)",
        "No總量(shares)",
        "總金額($)",
        "平均價格($)",
    ]
    col_widths_s = [8, 8, 8, 8, 16, 16, 14, 14]
    for col_idx, h in enumerate(stat_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(fill_type="solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths_s[
            col_idx - 1
        ]
    ws.row_dimensions[1].height = 28

    by_party = defaultdict(list)
    for t in all_trades:
        by_party[t["party"]].append(t)

    for row_idx, party in enumerate(["KMT", "DPP", "TPP"], 2):
        trades = by_party[party]
        buys = [t for t in trades if t["side"] == "BUY"]
        sells = [t for t in trades if t["side"] == "SELL"]
        yes_v = sum(t["size"] for t in trades if t["outcome"] == "Yes")
        no_v = sum(t["size"] for t in trades if t["outcome"] == "No")
        total = sum(t["total"] for t in trades)
        avg_p = sum(t["price"] for t in trades) / len(trades) if trades else 0

        row_data = [
            party,
            len(trades),
            len(buys),
            len(sells),
            round(yes_v, 2),
            round(no_v, 2),
            round(total, 2),
            round(avg_p, 4),
        ]
        bg, fg = PARTY_COLORS[party]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            cell.fill = PatternFill(
                fill_type="solid", fgColor=bg if col_idx == 1 else "F2F2F2"
            )
            cell.font = Font(
                bold=(col_idx == 1), color=fg if col_idx == 1 else "000000", size=11
            )

    # Monthly summary
    ws.cell(row=6, column=1, value="--- 月份分布 ---").font = Font(bold=True, size=11)
    monthly_h = ["月份", "總筆數", "KMT", "DPP", "TPP"]
    for col_idx, h in enumerate(monthly_h, 1):
        cell = ws.cell(row=7, column=col_idx, value=h)
        cell.fill = PatternFill(fill_type="solid", fgColor="2E4057")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r_off, mk in enumerate(monthly_keys, 8):
        trades_m = monthly[mk]
        kmt_n = sum(1 for t in trades_m if t["party"] == "KMT")
        dpp_n = sum(1 for t in trades_m if t["party"] == "DPP")
        tpp_n = sum(1 for t in trades_m if t["party"] == "TPP")
        for col_idx, val in enumerate([mk, len(trades_m), kmt_n, dpp_n, tpp_n], 1):
            cell = ws.cell(row=r_off, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            cell.fill = PatternFill(fill_type="solid", fgColor="F9F9F9")
            cell.font = Font(size=10)


# ── Step 1: Fetch all Polymarket trades ──────────────────────────────────────
print("=" * 55)
print("Step 1: Polymarket trades API")
print("=" * 55)

all_flat = []
for party, cid in MARKETS.items():
    trades = fetch_poly_trades(cid, party)
    for t in trades:
        all_flat.append(
            {
                "timestamp": t.get("timestamp", 0),
                "party": party,
                "side": t.get("side", ""),
                "outcome": t.get("outcome", ""),
                "size": round(float(t.get("size", 0)), 6),
                "price": round(float(t.get("price", 0)), 4),
                "total": round(float(t.get("size", 0)) * float(t.get("price", 0)), 4),
                "name": t.get("name", "") or t.get("pseudonym", "") or "",
                "proxyWallet": t.get("proxyWallet", ""),
                "transactionHash": t.get("transactionHash", ""),
            }
        )

all_flat.sort(key=lambda x: x["timestamp"])
print(f"  合計: {len(all_flat)}筆")

# Group by month
monthly = defaultdict(list)
for t in all_flat:
    if t["timestamp"] > 0:
        mk = month_key(t["timestamp"])
        monthly[mk].append(t)

monthly_keys = sorted(monthly.keys())
print(f"  月份: {monthly_keys[0]} ～ {monthly_keys[-1]}")
for mk in monthly_keys:
    print(f"    {mk}: {len(monthly[mk])}筆")

# ── Step 2: Build Excel ───────────────────────────────────────────────────────
print()
print("=" * 55)
print("Step 2: 建立 Excel")
print("=" * 55)

wb = Workbook()
border = thin_border()

ws_all = wb.active
ws_all.title = "總表"
write_header(ws_all, border)
write_rows(ws_all, all_flat, border)
print(f"  總表: {len(all_flat)}筆 OK")

for mk in monthly_keys:
    ws_m = wb.create_sheet(title=mk)
    write_header(ws_m, border)
    write_rows(ws_m, monthly[mk], border)
    print(f"  {mk}: {len(monthly[mk])}筆 OK")

write_summary_sheet(wb, all_flat, monthly_keys, monthly, border)
print("  統計摘要 OK")

output = "E:\\polymarket選舉賭博\\2026台灣地方選舉_全部交易.xlsx"
wb.save(output)
print(f"\nExcel 儲存至: {output}")
print(f"總計: {len(all_flat)} 筆（Polymarket API，經確認為完整資料集）")
print(
    "說明：Blockscout 上額外 87 筆均為 split/merge/內部操作，非方向性交易，已確認排除。"
)
