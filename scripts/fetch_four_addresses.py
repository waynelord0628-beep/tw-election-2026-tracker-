import urllib.request
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone

# ── 目標地址 ──────────────────────────────────────────────
TARGET_ADDRS = {
    "0xd8dd45139269031b16a54717cabad4af6a3980d6",
    "0x0d5ee5c536ccd78d179487b3d7e43ae4304d5c24",
    "0xfde3a53d58320a3db74dbe1092979c401e35719a",
    "0x426227d4a9c4ad5a1aae7f2706238f2154b9abaa",
}

# ── 三個市場 conditionId ───────────────────────────────────
MARKETS = {
    "KMT": "0xc0f076bc4d90a44df34a729277e9d1f294f0cb60d2c3b1b3800908b1e15b923b",
    "DPP": "0xea3b1d0099085f43a3098b3e1fbcbe62284ce1bda99384b3d46b82ff202ac016",
    "TPP": "0x68fc8d466ddc10a1ae37b52642f36b93e413cf98ba5fe3947c242a9a727b2e94",
}

# ── 備註資料（鏈上調查結果）───────────────────────────────
# key: transactionHash
NOTES = {
    # ChiangWan-an SELL 80 — 鏈上確認代幣流向
    "0x789c588349aadb4724642ee165a25d4dd79231909a03d757cd2c7edb471fa63c": "鏈上確認：直接轉出 80 KMT-Yes → NegRiskExchange，收到 68.8 USDC",
    # ChiangWan-an SELL 20 — 代幣來源不明
    "0xbaa910668e8413051767f452e0750debe9d7f03bd578681bbccbdbd3270398e0": "代幣來源不明：無 ERC-1155 Transfer 紀錄流入，疑為 negRisk adapter 內部持倉",
    # Kuomintang SELL 100
    "0x3d3c6abf90de71efc9775134c58afaf6f55c8d18d6e1fdfc8aa03f1bb380ad1a": "對應下方補錄的 BUY（2025-12-26）；先買後賣，獲利約 $3",
}

# ── 手動補錄的交易（Polymarket API 未收錄）────────────────
# 格式：[時間UTC, 政黨, 方向, 標的, 數量, 價格, 總金額, 名稱, 地址, Hash, 備註]
MANUAL_ROWS = [
    [
        "2025-12-26 21:47:31",  # 時間
        "KMT",  # 政黨
        "BUY",  # 方向
        "Yes",  # 標的
        100.0,  # 數量
        0.84,  # 價格
        84.0,  # 總金額
        "Kuomintang",  # 名稱
        "0x426227d4a9c4ad5a1aae7f2706238f2154b9abaa",  # 地址
        "0x23345629edb3f8ca6555eff4f1c88cd35ae83ec686d047132792768506161bd4",  # Hash
        "【補錄·鏈上調查】Polymarket API 未收錄此筆；鏈上確認：匯入 84 USDC → 收到 100 KMT-Yes（matchOrders）",
    ],
]


def http_get(url):
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=20) as r:
        return json.loads(r.read())


def fetch_market_trades(party, cond_id):
    trades = []
    offset = 0
    limit = 500
    while True:
        url = f"https://data-api.polymarket.com/trades?market={cond_id}&limit={limit}&offset={offset}"
        try:
            data = http_get(url)
        except Exception as e:
            print(f"  [!] {party} offset={offset}: {e}")
            break
        if not data:
            break
        for t in data:
            t["_party"] = party
            t["_conditionId"] = cond_id
        trades.extend(data)
        print(f"  {party}: offset={offset}, got {len(data)}, total={len(trades)}")
        if len(data) < limit:
            break
        offset += limit
    return trades


def ts_to_utc(ts):
    if ts is None:
        return ""
    try:
        dt = datetime.fromtimestamp(int(ts), tz=timezone.utc)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except:
        return str(ts)


# ── 抓取全部市場交易並過濾四地址 ─────────────────────────
print("抓取市場交易中...")
all_trades = []
for party, cid in MARKETS.items():
    market_trades = fetch_market_trades(party, cid)
    filtered = [
        t for t in market_trades if t.get("proxyWallet", "").lower() in TARGET_ADDRS
    ]
    print(f"  {party}: 市場總計 {len(market_trades)} 筆，四地址命中 {len(filtered)} 筆")
    all_trades.extend(filtered)

print(f"\n四地址合計 {len(all_trades)} 筆交易（不含補錄）")

# ── 整理 rows ─────────────────────────────────────────────
HEADERS = [
    "時間 (UTC)",
    "政黨",
    "方向",
    "標的 (Yes/No)",
    "數量 (Shares)",
    "價格 ($)",
    "總金額 ($)",
    "交易者名稱",
    "錢包地址",
    "交易 Hash",
    "備註",
]


def make_row(t):
    party = t.get("_party", "?")
    side = t.get("side", "")
    outcome = t.get("outcome", "")
    shares = float(t.get("size", 0) or 0)
    price = float(t.get("price", 0) or 0)
    total = round(shares * price, 4)
    ts = ts_to_utc(t.get("timestamp"))
    name = t.get("name") or t.get("pseudonym") or t.get("proxyWallet", "")[:12]
    addr = (t.get("proxyWallet") or "").lower()
    tx = t.get("transactionHash") or ""
    note = NOTES.get(tx, "")
    return [ts, party, side, outcome, shares, price, total, name, addr, tx, note]


rows = [make_row(t) for t in all_trades]
rows.extend(MANUAL_ROWS)
rows.sort(key=lambda r: r[0])

# ── Excel 樣式 ────────────────────────────────────────────
COLOR_HEADER = "1F4E79"
COLOR_PARTY = {"KMT": "D6E4F0", "DPP": "FDECEA", "TPP": "E8F5E9"}
COLOR_ADDR = {
    "0xd8dd45139269031b16a54717cabad4af6a3980d6": "FFF9C4",
    "0x0d5ee5c536ccd78d179487b3d7e43ae4304d5c24": "F3E5F5",
    "0xfde3a53d58320a3db74dbe1092979c401e35719a": "E0F7FA",
    "0x426227d4a9c4ad5a1aae7f2706238f2154b9abaa": "FBE9E7",
}
COLOR_BUY = "D5E8D4"  # 綠：BUY
COLOR_SELL = "FFE6CC"  # 橘：SELL
COLOR_MANUAL = "FFF2CC"  # 黃：補錄（鏈上調查）

thin = Side(style="thin", color="CCCCCC")

COL_WIDTHS = [20, 8, 8, 12, 14, 10, 12, 18, 46, 68, 70]


def hdr_style(ws):
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def row_color(row_data, use_addr_color=True):
    """決定行顏色：補錄 > BUY/SELL > 地址色"""
    note = row_data[10] if len(row_data) > 10 else ""
    side = row_data[2]
    addr = row_data[8]
    if "補錄" in str(note):
        return COLOR_MANUAL
    if use_addr_color:
        return COLOR_ADDR.get(addr, COLOR_BUY if side == "BUY" else COLOR_SELL)
    return COLOR_BUY if side == "BUY" else COLOR_SELL


def apply_row(ws_row, color):
    bdr = Border(top=thin, left=thin, right=thin, bottom=thin)
    for i, cell in enumerate(ws_row):
        cell.fill = PatternFill("solid", fgColor=color)
        cell.border = bdr
        # 備註欄自動換行
        if cell.column == 11:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        else:
            cell.alignment = Alignment(vertical="center")


wb = openpyxl.Workbook()

# ── Sheet 1: 全部交易 ──────────────────────────────────────
ws = wb.active
ws.title = "全部交易"
hdr_style(ws)
for row_data in rows:
    ws.append(row_data)
for r in ws.iter_rows(min_row=2, max_row=ws.max_row):
    rd = [c.value for c in r]
    apply_row(r, row_color(rd, use_addr_color=True))

# ── Sheet 2-N: 每個地址 ───────────────────────────────────
addr_names = {}
for t in all_trades:
    addr = (t.get("proxyWallet") or "").lower()
    name = (t.get("name") or t.get("pseudonym") or "").strip()
    if name and addr:
        addr_names[addr] = name
# 補錄 manual rows 的名稱
for mr in MANUAL_ROWS:
    addr_names[mr[8]] = mr[7]

for addr in sorted(TARGET_ADDRS):
    name = addr_names.get(addr, addr[:8])
    safe = (
        name[:28]
        .replace("/", "-")
        .replace("\\", "-")
        .replace("?", "")
        .replace("*", "")
        .replace("[", "")
        .replace("]", "")
        .replace(":", "")
    )
    ws2 = wb.create_sheet(title=safe or addr[:8])
    hdr_style(ws2)

    addr_rows = [r for r in rows if r[8] == addr]
    for row_data in addr_rows:
        ws2.append(row_data)

    for r in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        rd = [c.value for c in r]
        apply_row(r, row_color(rd, use_addr_color=False))

    # 若只有 SELL 或只有 BUY，加上提示列
    sides = [r[2] for r in addr_rows]
    has_buy = any(s == "BUY" for s in sides)
    has_sell = any(s == "SELL" for s in sides)

    note_row = ws2.max_row + 2
    if has_buy and not has_sell:
        msg = "⚠ 此地址目前只有 BUY 紀錄（尚未賣出，持倉中）"
        c = ws2.cell(note_row, 1, msg)
        c.font = Font(bold=True, color="0070C0", size=11)
    elif has_sell and not has_buy:
        msg = "⚠ 此地址只有 SELL 紀錄（BUY 未被 Polymarket API 收錄，詳見「代幣來源調查」分頁）"
        c = ws2.cell(note_row, 1, msg)
        c.font = Font(bold=True, color="C00000", size=11)
        ws2.merge_cells(f"A{note_row}:K{note_row}")

    # 右側摘要
    col = 13
    ws2.cell(2, col, "=== 摘要 ===").font = Font(bold=True, size=12)
    buy_rows = [r for r in addr_rows if r[2] == "BUY"]
    sell_rows = [r for r in addr_rows if r[2] == "SELL"]
    manual_buy_rows = [r for r in buy_rows if "補錄" in str(r[10])]
    summaries = [
        ("總交易筆數", len(addr_rows)),
        ("　其中補錄（鏈上）", len(manual_buy_rows)),
        ("買入筆數", len(buy_rows)),
        ("賣出筆數", len(sell_rows)),
        ("買入總金額 ($)", round(sum(r[6] for r in buy_rows), 2)),
        ("賣出總金額 ($)", round(sum(r[6] for r in sell_rows), 2)),
        (
            "損益估算 ($)",
            round(sum(r[6] for r in sell_rows) - sum(r[6] for r in buy_rows), 2),
        ),
        ("KMT 筆數", sum(1 for r in addr_rows if r[1] == "KMT")),
        ("DPP 筆數", sum(1 for r in addr_rows if r[1] == "DPP")),
        ("TPP 筆數", sum(1 for r in addr_rows if r[1] == "TPP")),
    ]
    for i, (label, val) in enumerate(summaries, 3):
        ws2.cell(i, col, label).font = Font(bold=True)
        ws2.cell(i, col + 1, val)

# ── 代幣來源調查 sheet ────────────────────────────────────
INV_HEADERS = [
    "錢包地址",
    "名稱",
    "市場",
    "操作類型",
    "時間 (UTC)",
    "數量 (Shares)",
    "價格 ($)",
    "金額 (USDC)",
    "交易 Hash",
    "說明",
]
INV_COL_WIDTHS = [46, 16, 8, 22, 22, 14, 10, 12, 68, 80]
INV_DATA = [
    (
        "0x426227d4a9c4ad5a1aae7f2706238f2154b9abaa",
        "Kuomintang",
        "KMT",
        "BUY（補錄·鏈上）",
        "2025-12-26 21:47:31",
        100,
        0.84,
        84.0,
        "0x23345629edb3f8ca6555eff4f1c88cd35ae83ec686d047132792768506161bd4",
        "Polymarket API 未收錄；鏈上確認匯入 84 USDC → 收到 100 KMT-Yes（matchOrders）",
    ),
    (
        "0x426227d4a9c4ad5a1aae7f2706238f2154b9abaa",
        "Kuomintang",
        "KMT",
        "SELL（Polymarket）",
        "2026-02-07 01:20:11",
        100,
        0.87,
        87.0,
        "0x3d3c6abf90de71efc9775134c58afaf6f55c8d18d6e1fdfc8aa03f1bb380ad1a",
        "賣出先前 BUY 的 100 股；獲利約 $3（$87 - $84）",
    ),
    (
        "0x0d5ee5c536ccd78d179487b3d7e43ae4304d5c24",
        "ChiangWan-an",
        "KMT",
        "BUY？（來源不明）",
        "不明",
        100,
        "?",
        "?",
        "無",
        "無 ERC-1155 Transfer 入帳紀錄，推測透過 negRisk adapter 內部持倉（買其他候選人 No ≈ 持有 KMT Yes）",
    ),
    (
        "0x0d5ee5c536ccd78d179487b3d7e43ae4304d5c24",
        "ChiangWan-an",
        "KMT",
        "SELL（Polymarket）",
        "2026-02-07 01:20:33",
        20,
        0.87,
        17.4,
        "0xbaa910668e8413051767f452e0750debe9d7f03bd578681bbccbdbd3270398e0",
        "賣出 20 股；代幣來源不明",
    ),
    (
        "0x0d5ee5c536ccd78d179487b3d7e43ae4304d5c24",
        "ChiangWan-an",
        "KMT",
        "SELL（Polymarket）",
        "2026-04-10 12:37:13",
        80,
        0.86,
        68.8,
        "0x789c588349aadb4724642ee165a25d4dd79231909a03d757cd2c7edb471fa63c",
        "鏈上確認：直接轉出 80 KMT-Yes → NegRiskExchange，收到 68.8 USDC",
    ),
]
CONCLUSIONS = [
    "【Kuomintang】有完整 BUY（2025-12-26, $0.84）→ SELL（2026-02-07, $0.87）流程，獲利約 $3。",
    "  BUY 未被 Polymarket trades API 收錄（直接合約操作），故原始資料只顯示 SELL。",
    "",
    "【ChiangWan-an】代幣來源無 ERC-1155 Transfer 紀錄，疑為 negRisk adapter 內部持倉。",
    "  SELL 20 股（2026-02-07）與 SELL 80 股（2026-04-10）均已確認。",
    "",
    "【共同規律】兩地址的 BUY 操作均未出現在 Polymarket trades API，",
    "  可能原因：直接合約交互 / negRisk adapter 操作 / 批次撮合未被單獨記錄。",
]

ws_inv = wb.create_sheet("代幣來源調查")
for col, h in enumerate(INV_HEADERS, 1):
    c = ws_inv.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
ws_inv.row_dimensions[1].height = 30

for ri, row in enumerate(INV_DATA, 2):
    action = row[3]
    if "BUY" in action and "補錄" in action:
        bg = "FFF2CC"
    elif "BUY" in action:
        bg = "D5E8D4"
    elif "SELL" in action:
        bg = "FFE6CC"
    else:
        bg = "FFF9C4"
    for ci, val in enumerate(row, 1):
        c = ws_inv.cell(row=ri, column=ci, value=val)
        c.fill = PatternFill("solid", fgColor=bg)
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        c.alignment = Alignment(vertical="top", wrap_text=(ci in (1, 10)))

note_row = len(INV_DATA) + 3
ws_inv.cell(note_row, 1, "【調查結論】").font = Font(bold=True, size=12, color="CC0000")
ws_inv.merge_cells(f"A{note_row}:J{note_row}")
for i, line in enumerate(CONCLUSIONS, note_row + 1):
    c = ws_inv.cell(i, 1, line)
    c.alignment = Alignment(wrap_text=True)
    c.font = (
        Font(color="555555")
        if line.startswith("  ")
        else Font(bold=(line != ""), color="333333")
    )
    ws_inv.merge_cells(f"A{i}:J{i}")

for i, w in enumerate(INV_COL_WIDTHS, 1):
    ws_inv.column_dimensions[get_column_letter(i)].width = w
ws_inv.freeze_panes = "A2"

# ── 欄寬 / 凍結 / 篩選 ────────────────────────────────────
for sheet in wb.worksheets:
    for i, w in enumerate(COL_WIDTHS, 1):
        sheet.column_dimensions[get_column_letter(i)].width = w
    sheet.column_dimensions[get_column_letter(13)].width = 18
    sheet.column_dimensions[get_column_letter(14)].width = 14
    sheet.row_dimensions[1].height = 30
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = f"A1:K{sheet.max_row}"

out = r"E:\polymarket選舉賭博\四地址台灣選舉交易.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Total rows (含補錄): {len(rows)}")
