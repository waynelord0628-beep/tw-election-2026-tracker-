"""
build_complete_excel.py
從 Blockscout 重建完整交易明細（含 LP/AMM 交易）
─────────────────────────────────────────────────────────────────
策略：
  1. 對每個 CTF token，從 Blockscout 抓全部 transfers
  2. 過濾掉合約間轉帳（CTF / Exchange / Adapter / zero）
  3. 排除 split/merge（同一 tx 中 Yes+No 對同一用戶同方向移動）
  4. 價格：API 有的用 API price；沒有的用 Blockscout USDC 流向計算
  5. 輸出完整 Excel（與舊版欄位相容）
"""

import requests, time, json, os
from datetime import datetime, timezone, timedelta
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TZ8 = timezone(timedelta(hours=8))

# ─── 合約地址 ───────────────────────────────────────────────────────
CTF = "0x4d97dcd97ec945f40cf65f87097ace5ea0476045"
EXCHANGE = "0xc5d563a36ae78145c45a50134d48a1215220f80a"  # NegRiskCtfExchange
ADAPTER = "0xd91e80cf2e7be2e162c6513ced06f1dd0da35296"  # NegRisk Adapter
ZERO = "0x0000000000000000000000000000000000000000"
USDC = "0x2791bca1f2de4661ed88a30c99a7a9449aa84174"
KNOWN_CONTRACTS = {CTF, EXCHANGE, ADAPTER, ZERO}

# ─── 市場 / Token ───────────────────────────────────────────────────
MARKETS = {
    "KMT": "0xc0f076bc4d90a44df34a729277e9d1f294f0cb60d2c3b1b3800908b1e15b923b",
    "DPP": "0xea3b1d0099085f43a3098b3e1fbcbe62284ce1bda99384b3d46b82ff202ac016",
    "TPP": "0x68fc8d466ddc10a1ae37b52642f36b93e413cf98ba5fe3947c242a9a727b2e94",
}

TOKENS = {
    "KMT_Yes": (
        "85632914518786177256583369552125280053108667306405854845853340618248288927460",
        "KMT",
        "Yes",
    ),
    "KMT_No": (
        "4696955573632845407532815267539406678302911508204032661527405293140196109387",
        "KMT",
        "No",
    ),
    "DPP_Yes": (
        "13628189982642424912108657221169198338993179248246381972030640500448717195916",
        "DPP",
        "Yes",
    ),
    "DPP_No": (
        "91004506882941445266754771479824617369805789899332711132070603219216406556613",
        "DPP",
        "No",
    ),
    "TPP_Yes": (
        "14999500579901383072635205035227864886528710236540822730141548371372688859422",
        "TPP",
        "Yes",
    ),
    "TPP_No": (
        "16222840603445450947154718759167300491302153317593739623696847197718420087623",
        "TPP",
        "No",
    ),
}

PARTY_COLORS = {
    "KMT": ("002868", "FFFFFF"),
    "DPP": ("1B9431", "FFFFFF"),
    "TPP": ("28C8C8", "000000"),
}

HEADERS = [
    "時間 (UTC+8)",
    "政黨",
    "方向",
    "標的 (Yes/No)",
    "數量 (Shares)",
    "價格 ($)",
    "總金額 ($)",
    "交易者名稱",
    "錢包地址",
    "交易 Hash",
    "備註",
]
COL_WIDTHS = [22, 6, 6, 14, 16, 10, 12, 20, 44, 68, 16]

# ─── Step 1: 從 API 收集所有 KMT/DPP/TPP 交易（用於價格查詢）───────
print("=== Step 1: Fetch Polymarket API trades (for price reference) ===")
api_trades = []  # all trades
api_price_map = {}  # (txhash, token_id) -> price
for party, cond in MARKETS.items():
    offset = 0
    count = 0
    while True:
        r = requests.get(
            f"https://data-api.polymarket.com/trades?market={cond}&limit=500&offset={offset}",
            timeout=30,
        )
        batch = r.json()
        if not batch:
            break
        for t in batch:
            txh = t.get("transactionHash", "").lower()
            outcome = t.get("outcome", "")
            token_key = f"{party}_{outcome}"
            if token_key in TOKENS:
                tid = TOKENS[token_key][0]
                price = float(t.get("price", 0))
                if txh and price > 0:
                    api_price_map[(txh, tid)] = price
        api_trades.extend(batch)
        count += len(batch)
        if len(batch) < 500:
            break
        offset += 500
        time.sleep(0.2)
    print(f"  {party}: {count} trades")

print(f"API total: {len(api_trades)} trades, {len(api_price_map)} price records")

# ─── Step 2: 從 Blockscout 抓所有 token transfers ──────────────────
print("\n=== Step 2: Fetch Blockscout token transfers ===")

# transfers: list of {token_id, token_key, party, outcome, txhash, timestamp, wallet, direction, shares}
all_transfers = []

CTF_ORIGINAL = "0x4D97DCd97eC945f40cF65F87097ACe5EA0476045"

for token_key, (tid, party, outcome) in TOKENS.items():
    print(f"  {token_key}...", end="", flush=True)
    url = f"https://polygon.blockscout.com/api/v2/tokens/{CTF_ORIGINAL}/instances/{tid}/transfers"
    page_params = None
    batch_count = 0
    while True:
        try:
            r = requests.get(url, params=page_params, timeout=30)
            if r.status_code != 200:
                print(f" [HTTP {r.status_code}]", end="")
                break
            d = r.json()
        except Exception as e:
            print(f" [ERR: {e}]", end="")
            break
        items = d.get("items", [])
        for item in items:
            frm = item.get("from", {}).get("hash", "").lower()
            to_ = item.get("to", {}).get("hash", "").lower()
            txh = item.get("transaction_hash", "").lower()
            ts = item.get("timestamp", "")
            val = int(item.get("total", {}).get("value", 0)) / 1e6

            # 找出「用戶」錢包（非合約）
            # from = 合約, to = 用戶 → BUY
            # from = 用戶, to = 合約 → SELL
            # 雙方都是合約 → 跳過
            frm_is_contract = frm in KNOWN_CONTRACTS
            to_is_contract = to_ in KNOWN_CONTRACTS

            if frm_is_contract and to_is_contract:
                continue  # contract-to-contract, skip

            if not frm_is_contract and not to_is_contract:
                # user-to-user（罕見，記錄兩邊）
                all_transfers.append(
                    {
                        "token_key": token_key,
                        "token_id": tid,
                        "party": party,
                        "outcome": outcome,
                        "txhash": txh,
                        "timestamp": ts,
                        "wallet": to_,
                        "direction": "BUY",
                        "shares": val,
                    }
                )
                all_transfers.append(
                    {
                        "token_key": token_key,
                        "token_id": tid,
                        "party": party,
                        "outcome": outcome,
                        "txhash": txh,
                        "timestamp": ts,
                        "wallet": frm,
                        "direction": "SELL",
                        "shares": val,
                    }
                )
                batch_count += 2
                continue

            if frm_is_contract:
                wallet = to_
                direction = "BUY"
            else:
                wallet = frm
                direction = "SELL"

            all_transfers.append(
                {
                    "token_key": token_key,
                    "token_id": tid,
                    "party": party,
                    "outcome": outcome,
                    "txhash": txh,
                    "timestamp": ts,
                    "wallet": wallet,
                    "direction": direction,
                    "shares": val,
                }
            )
            batch_count += 1

        nxt = d.get("next_page_params")
        if not nxt:
            break
        page_params = nxt
        time.sleep(0.3)
    print(f" {batch_count} user transfers")

print(f"Total user transfers: {len(all_transfers)}")

# ─── Step 3: 偵測並排除 split/merge ────────────────────────────────
print("\n=== Step 3: Detect split/merge transactions ===")

# 同一 tx 中，同一 wallet，同時有 Yes + No 同方向移動 → split/merge
tx_wallet_tokens = defaultdict(lambda: defaultdict(set))
for t in all_transfers:
    key = (t["txhash"], t["wallet"], t["direction"])
    tx_wallet_tokens[key]["outcomes"].add(t["outcome"])
    tx_wallet_tokens[key]["parties"].add(t["party"])

split_merge_keys = set()
for (txh, wallet, direction), data in tx_wallet_tokens.items():
    if "Yes" in data["outcomes"] and "No" in data["outcomes"]:
        split_merge_keys.add((txh, wallet))

print(f"Split/merge (txhash, wallet) pairs: {len(split_merge_keys)}")

clean_transfers = [
    t for t in all_transfers if (t["txhash"], t["wallet"]) not in split_merge_keys
]
print(f"Clean transfers after removing split/merge: {len(clean_transfers)}")

# ─── Step 4: 去重 ──────────────────────────────────────────────────
print("\n=== Step 4: Dedup ===")

# 4a. 同一 (txhash, wallet, token_key) → 只保留 shares 最大的那筆
dedup = {}
for t in clean_transfers:
    key = (t["txhash"], t["wallet"], t["token_key"])
    if key not in dedup or t["shares"] > dedup[key]["shares"]:
        dedup[key] = t
stage1 = list(dedup.values())
print(f"  After wallet-level dedup: {len(stage1)}")

# 4b. CLOB 對手方標記：
#   同一 (txhash, token_key) 若同時存在 BUY 和 SELL（不同 wallet），
#   在 SELL 那方的 note 欄標記 "CLOB對手方"，BUY 那方留空。
#   純出場 SELL（沒有對應 BUY）note 留空。
from collections import defaultdict as _dd

tx_token_dirs = _dd(set)  # (txhash, token_key) -> {directions}
for t in stage1:
    tx_token_dirs[(t["txhash"], t["token_key"])].add(t["direction"])

counterparty_sell_keys = set()
for (txh, tk), dirs in tx_token_dirs.items():
    if "BUY" in dirs and "SELL" in dirs:
        counterparty_sell_keys.add((txh, tk))

for t in stage1:
    if (t["txhash"], t["token_key"]) in counterparty_sell_keys and t[
        "direction"
    ] == "SELL":
        t["note"] = "CLOB對手方"
    else:
        t["note"] = ""

unique_transfers = stage1
clob_count = sum(1 for t in stage1 if t.get("note") == "CLOB對手方")
print(f"  標記 {clob_count} 筆 CLOB對手方（保留在資料中）")
print(f"After dedup: {len(unique_transfers)} unique trades")

# ─── Step 5: 補充價格 ───────────────────────────────────────────────
print("\n=== Step 5: Fill prices ===")

# 需要額外抓價格的 txhash set
need_price = []
has_price = 0
for t in unique_transfers:
    key = (t["txhash"], t["token_id"])
    if key in api_price_map:
        t["price"] = api_price_map[key]
        has_price += 1
    else:
        t["price"] = None
        need_price.append(t)

print(f"  From API: {has_price} trades have price")
print(f"  Need Blockscout price: {len(need_price)} trades")

# 對需要額外抓的，從 Blockscout tx token-transfers 找 USDC 流向
unique_need_txhash = list(set(t["txhash"] for t in need_price))
print(f"  Unique txhashes to query: {len(unique_need_txhash)}")

# 快取檔案
USDC_CACHE_FILE = r"E:\polymarket選舉賭博\usdc_cache.json"
if os.path.exists(USDC_CACHE_FILE):
    with open(USDC_CACHE_FILE) as f:
        tx_usdc_map = json.load(f)
    print(f"  Loaded cache: {len(tx_usdc_map)} txs")
else:
    tx_usdc_map = {}

to_fetch = [txh for txh in unique_need_txhash if txh not in tx_usdc_map]
print(
    f"  Need to fetch: {len(to_fetch)} txs (cached: {len(unique_need_txhash) - len(to_fetch)})"
)


def fetch_usdc_for_tx(txh):
    try:
        r = requests.get(
            f"https://polygon.blockscout.com/api/v2/transactions/{txh}/token-transfers",
            timeout=15,
        )
        if r.status_code == 200:
            items = r.json().get("items", [])
            usdc_flows = {}
            for item in items:
                sym = (item.get("token", {}).get("symbol") or "").upper()
                if sym not in ("USDC", "USDC.E"):
                    continue
                frm2 = item.get("from", {}).get("hash", "").lower()
                to_2 = item.get("to", {}).get("hash", "").lower()
                v = int(item.get("total", {}).get("value", 0)) / 1e6
                if to_2 not in KNOWN_CONTRACTS:
                    usdc_flows[to_2] = usdc_flows.get(to_2, 0) + v
                if frm2 not in KNOWN_CONTRACTS:
                    usdc_flows[frm2] = usdc_flows.get(frm2, 0) - v
            return txh, usdc_flows
    except Exception:
        pass
    return txh, {}


if to_fetch:
    done = 0
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(fetch_usdc_for_tx, txh): txh for txh in to_fetch}
        for future in as_completed(futures):
            txh, flows = future.result()
            tx_usdc_map[txh] = flows
            done += 1
            if done % 50 == 0:
                print(f"    fetched {done}/{len(to_fetch)}...", flush=True)
    # 存快取
    with open(USDC_CACHE_FILE, "w") as f:
        json.dump(tx_usdc_map, f)
    print(f"  Saved cache: {len(tx_usdc_map)} txs")

# 對每筆 need_price，用 USDC 流向計算價格
priced_from_chain = 0
unpriced = 0
for t in need_price:
    txh = t["txhash"]
    wallet = t["wallet"]
    usdc_flows = tx_usdc_map.get(txh, {})
    usdc = usdc_flows.get(wallet, 0)
    shares = t["shares"]
    if shares > 0 and abs(usdc) > 0:
        price = abs(usdc) / shares
        # sanity check: price should be between 0 and 1
        if 0 < price <= 1.0:
            t["price"] = round(price, 6)
            priced_from_chain += 1
        else:
            t["price"] = None
            unpriced += 1
    else:
        t["price"] = None
        unpriced += 1

print(f"  Priced from chain: {priced_from_chain}")
print(f"  Still unpriced (will use 0): {unpriced}")

# ─── Step 6: 補充交易者名稱 ──────────────────────────────────────────
KNOWN_WALLETS = {
    "0xd8dd45139269031b16a54717cabad4af6a3980d6": "調查目標-1",
    "0x0d5ee5c536ccd78d179487b3d7e43ae4304d5c24": "ChiangWan-an",
    "0xfde3a53d58320a3db74dbe1092979c401e35719a": "jamieamoy",
    "0x426227d4a9c4ad5a1aae7f2706238f2154b9abaa": "Kuomintang",
    "0xc8ab97a9089a9ff7e6ef0688e6e591a066946418": "ArmageddonRewardsBilly",
    "0xc2358d03312b05b244bde5286dee03bc60ac99f8": "Anon-0xCE",
    "0x0362bb926368b144e0ff98f6828a251e4cb6449e": "varch01",
    "0x0cd7bea497efb9220105858617f0cd660d0a78e0": "SirJason",
    "0x06d248d4f372601d24192284bff919a2c05dfb27": "cheesymm",
    "0xc25120b27e01031b2122f74488dcdb077a78b9c3": "TTbilly",
}

# 從 API trades 補充名稱
for t_api in api_trades:
    w = t_api.get("proxyWallet", "").lower()
    name = t_api.get("name", "") or t_api.get("pseudonym", "")
    if w and name and w not in KNOWN_WALLETS:
        KNOWN_WALLETS[w] = name

print(f"\nKnown wallets: {len(KNOWN_WALLETS)}")


def get_name(wallet):
    return KNOWN_WALLETS.get(wallet.lower(), "")


# ─── Step 7: 建立最終行列表 ──────────────────────────────────────────
print("\n=== Step 7: Build rows ===")

rows = []
for t in unique_transfers:
    ts_str = t.get("timestamp", "")
    try:
        dt_utc = datetime.fromisoformat(ts_str.replace("Z", "+00:00"))
        dt_local = dt_utc.astimezone(TZ8)
        time_str = dt_local.strftime("%Y-%m-%d %H:%M:%S")
    except:
        time_str = ts_str

    party = t["party"]
    direction = t["direction"]
    outcome = t["outcome"]
    shares = round(t["shares"], 4)
    price = t.get("price") or 0
    total = round(shares * price, 4) if price else 0
    wallet = t["wallet"]
    name = get_name(wallet)
    txhash = t["txhash"]

    rows.append(
        [
            time_str,
            party,
            direction,
            outcome,
            shares,
            round(price, 6),
            total,
            name,
            wallet,
            txhash,
            t.get("note", ""),
        ]
    )

# Sort by time
rows.sort(key=lambda r: r[0])
print(f"Total rows: {len(rows)}")

# ─── Step 8: 輸出 Excel ──────────────────────────────────────────────
print("\n=== Step 8: Write Excel ===")

wb = Workbook()


def make_sheet(ws, sheet_rows, title=""):
    ws.freeze_panes = "A2"
    # Header
    for col_idx, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    for row_idx, row in enumerate(sheet_rows, 2):
        party = row[1]
        bg, fg = PARTY_COLORS.get(party, ("FFFFFF", "000000"))
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 2:  # 政黨列上色
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(color=fg, bold=True)
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 3:  # 方向
                color = "C6EFCE" if val == "BUY" else "FFCCCC"
                cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = Alignment(horizontal="center")


# 總表
ws_all = wb.active
ws_all.title = "總表"
make_sheet(ws_all, rows)

# 月份分頁
from collections import defaultdict as dd

months = dd(list)
for row in rows:
    ym = row[0][:7]  # "YYYY-MM"
    months[ym].append(row)

for ym in sorted(months.keys()):
    ws = wb.create_sheet(title=ym)
    make_sheet(ws, months[ym])

# 統計摘要
ws_stat = wb.create_sheet(title="統計摘要")
ws_stat["A1"] = "統計項目"
ws_stat["B1"] = "數值"
ws_stat["A1"].font = Font(bold=True)
ws_stat["B1"].font = Font(bold=True)

total_rows = len(rows)
party_counts = defaultdict(int)
party_vol = defaultdict(float)
for row in rows:
    party_counts[row[1]] += 1
    party_vol[row[1]] += row[6]

stats = [
    ("總交易筆數（含 LP/AMM）", total_rows),
    ("KMT 交易筆數", party_counts["KMT"]),
    ("DPP 交易筆數", party_counts["DPP"]),
    ("TPP 交易筆數", party_counts["TPP"]),
    ("KMT 總交易量 (USD)", round(party_vol["KMT"], 2)),
    ("DPP 總交易量 (USD)", round(party_vol["DPP"], 2)),
    ("TPP 總交易量 (USD)", round(party_vol["TPP"], 2)),
    ("資料來源", "Blockscout 鏈上 + Polymarket API（含 LP/AMM）"),
    ("舊版 CLOB-only 筆數", 964),
    ("新增 LP/AMM 交易筆數", total_rows - 964 if total_rows > 964 else "N/A"),
]
for i, (k, v) in enumerate(stats, 2):
    ws_stat.cell(row=i, column=1, value=k)
    ws_stat.cell(row=i, column=2, value=v)
ws_stat.column_dimensions["A"].width = 35
ws_stat.column_dimensions["B"].width = 50

# ─── 查詢分頁（實際資料 + VBA 搜尋）─────────────────────────────────
ws_q = wb.create_sheet(title="查詢")

BLUE_DARK = "1F3864"
YELLOW_BG = "FFFF99"
GRAY_BG = "F2F2F2"
REF_HDR_BG = "2E4057"  # 深藍灰，區分標題
REF_ALT_BG = "EEF2FF"

# ── 第一列：標題 ──
ws_q.merge_cells("A1:K1")
title_cell = ws_q["A1"]
title_cell.value = "使用者交易查詢"
title_cell.font = Font(bold=True, size=16, color="FFFFFF")
title_cell.fill = PatternFill("solid", fgColor=BLUE_DARK)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_q.row_dimensions[1].height = 30

# ── 第二列：快速參考表標題 ──
ws_q.merge_cells("A2:K2")
ref_title = ws_q["A2"]
ref_title.value = "▶  已知用戶快速參考（點名稱可複製到搜尋框）"
ref_title.font = Font(bold=True, size=11, color="FFFFFF")
ref_title.fill = PatternFill("solid", fgColor=REF_HDR_BG)
ref_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_q.row_dimensions[2].height = 20

# ── 第三列：參考表欄頭 ──
for col, label in [(1, "名稱"), (2, "錢包地址")]:
    c = ws_q.cell(row=3, column=col, value=label)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=REF_HDR_BG)
    c.alignment = Alignment(horizontal="center")
ws_q.row_dimensions[3].height = 18

FEATURED_WALLETS = [
    ("調查目標-1", "0xd8dd45139269031b16a54717cabad4af6a3980d6"),
    ("ChiangWan-an", "0x0d5ee5c536ccd78d179487b3d7e43ae4304d5c24"),
    ("jamieamoy", "0xfde3a53d58320a3db74dbe1092979c401e35719a"),
    ("Kuomintang", "0x426227d4a9c4ad5a1aae7f2706238f2154b9abaa"),
    ("ArmageddonRewardsBilly", "0xc8ab97a9089a9ff7e6ef0688e6e591a066946418"),
    ("Anon-0xCE", "0xc2358d03312b05b244bde5286dee03bc60ac99f8"),
    ("varch01", "0x0362bb926368b144e0ff98f6828a251e4cb6449e"),
    ("SirJason", "0x0cd7bea497efb9220105858617f0cd660d0a78e0"),
    ("cheesymm", "0x06d248d4f372601d24192284bff919a2c05dfb27"),
    ("TTbilly", "0xc25120b27e01031b2122f74488dcdb077a78b9c3"),
]
# 列 4–13
for i, (name, addr) in enumerate(FEATURED_WALLETS, 4):
    ws_q.row_dimensions[i].height = 17
    c_name = ws_q.cell(row=i, column=1, value=name)
    c_addr = ws_q.cell(row=i, column=2, value=addr)
    c_name.font = Font(bold=True, size=10)
    c_addr.font = Font(size=9, color="333333")
    if i % 2 == 0:
        for col in [1, 2]:
            ws_q.cell(row=i, column=col).fill = PatternFill("solid", fgColor=REF_ALT_BG)

ws_q.column_dimensions["A"].width = 26
ws_q.column_dimensions["B"].width = 46

# ── 列 14：分隔線 ──
ws_q.row_dimensions[14].height = 6

# ── 列 15：搜尋輸入 ──
ws_q["A15"] = "搜尋關鍵字："
ws_q["A15"].font = Font(bold=True, size=12)
ws_q["A15"].alignment = Alignment(horizontal="right", vertical="center")
ws_q.row_dimensions[15].height = 24

b15 = ws_q["B15"]
b15.fill = PatternFill("solid", fgColor=YELLOW_BG)
b15.font = Font(bold=True, size=12, color="1F3864")
b15.alignment = Alignment(horizontal="left", vertical="center")

ws_q["D15"] = "<- 輸入後按 Enter，結果自動更新"
ws_q["D15"].font = Font(italic=True, color="888888", size=10)

# ── 列 16：空白 ──
ws_q.row_dimensions[16].height = 4

# ── 列 17–18：統計 ──
stat_labels = [
    ("A17", "符合筆數"),
    ("D17", "BUY 總量 (Shares)"),
    ("F17", "SELL 總量 (Shares)"),
    ("H17", "淨持倉 (BUY-SELL)"),
    ("D18", "總花費 USD (BUY)"),
    ("F18", "總收入 USD (SELL)"),
    ("H18", "淨花費 USD"),
]
for addr, label in stat_labels:
    c = ws_q[addr]
    c.value = label
    c.font = Font(bold=True)
    c.fill = PatternFill("solid", fgColor=GRAY_BG)

for addr in ["B17", "E17", "G17", "I17", "E18", "G18", "I18"]:
    c = ws_q[addr]
    c.fill = PatternFill("solid", fgColor=GRAY_BG)
    c.font = Font(bold=True, color="CC0000")

# ── 列 19：分隔線 ──
ws_q.row_dimensions[19].height = 4

# ── 列 20：資料表頭 ──
for col_idx, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
    cell = ws_q.cell(row=20, column=col_idx, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=BLUE_DARK)
    cell.alignment = Alignment(horizontal="center")
    ws_q.column_dimensions[get_column_letter(col_idx)].width = w

# 凍結 A21（資料從 21 列開始）
ws_q.freeze_panes = "A21"

print("查詢分頁架構完成")

# ─── Step 9: 儲存 xlsx，再用 win32com 注入 VBA → 另存 xlsm ──────────
xlsx_path = r"E:\polymarket選舉賭博\_temp_build.xlsx"
xlsm_path = r"E:\polymarket選舉賭博\2026台灣地方選舉_完整交易_含LP.xlsm"

wb.save(xlsx_path)
print(f"Temp xlsx saved: {xlsx_path}")

# VBA 程式碼：Worksheet_Change 事件在 B3 變動時觸發搜尋
VBA_SHEET_CODE = r"""
Option Explicit

Private Sub Worksheet_Change(ByVal Target As Range)
    If Target.Address = "$B$15" Then
        Call SearchUser(Me.Range("B15").Value)
    End If
End Sub
"""

VBA_MODULE_CODE = r"""
Option Explicit

Sub ClearSearch()
    Dim wsQ As Worksheet
    Set wsQ = ThisWorkbook.Sheets(8)
    Application.EnableEvents = False
    wsQ.Range("B15").ClearContents
    Application.EnableEvents = True
    wsQ.Range("A21:K" & wsQ.Rows.Count).ClearContents
    wsQ.Range("A21:K" & wsQ.Rows.Count).Interior.ColorIndex = xlNone
    wsQ.Range("B17,E17,G17,I17,E18,G18,I18").Value = 0
End Sub

Sub SearchUser(keyword As String)
    Dim wsData As Worksheet
    Dim wsQ As Worksheet
    Dim dataRows As Long
    Dim i As Long, outRow As Long
    Dim keyword_lower As String
    Dim name_val As String, wallet_val As String
    Dim buyShares As Double, sellShares As Double
    Dim buyUSD As Double, sellUSD As Double
    Dim matchCount As Long
    Dim direction As String
    Dim shares As Double, usdVal As Double
    
    Set wsData = ThisWorkbook.Sheets("AllTrades")
    Set wsQ = ThisWorkbook.Sheets(8)  '  查詢分頁（第8個）
    
    ' 清空舊結果（第 21 列起）
    wsQ.Range("A21:K" & wsQ.Rows.Count).ClearContents
    wsQ.Range("A21:K" & wsQ.Rows.Count).Interior.ColorIndex = xlNone
    
    If Trim(keyword) = "" Then
        wsQ.Range("B17").Value = 0
        wsQ.Range("E17").Value = 0
        wsQ.Range("G17").Value = 0
        wsQ.Range("I17").Value = 0
        wsQ.Range("E18").Value = 0
        wsQ.Range("G18").Value = 0
        wsQ.Range("I18").Value = 0
        Exit Sub
    End If
    
    keyword_lower = LCase(keyword)
    dataRows = wsData.Cells(wsData.Rows.Count, 1).End(xlUp).Row
    outRow = 21
    matchCount = 0
    buyShares = 0 : sellShares = 0
    buyUSD = 0 : sellUSD = 0
    
    Application.ScreenUpdating = False
    
    For i = 2 To dataRows
        name_val   = LCase(CStr(wsData.Cells(i, 8).Value))
        wallet_val = LCase(CStr(wsData.Cells(i, 9).Value))
        
        If InStr(name_val, keyword_lower) > 0 Or InStr(wallet_val, keyword_lower) > 0 Then
            ' 複製整列（含備註欄）
            wsData.Rows(i).Copy wsQ.Rows(outRow)
            
            direction = CStr(wsData.Cells(i, 3).Value)
            shares    = CDbl(wsData.Cells(i, 5).Value)
            usdVal    = CDbl(wsData.Cells(i, 7).Value)
            
            If direction = "BUY" Then
                buyShares = buyShares + shares
                buyUSD = buyUSD + usdVal
            ElseIf direction = "SELL" Then
                sellShares = sellShares + shares
                sellUSD = sellUSD + usdVal
            End If
            
            outRow = outRow + 1
            matchCount = matchCount + 1
        End If
    Next i
    
    ' 更新統計
    wsQ.Range("B17").Value = matchCount
    wsQ.Range("E17").Value = Round(buyShares, 4)
    wsQ.Range("G17").Value = Round(sellShares, 4)
    wsQ.Range("I17").Value = Round(buyShares - sellShares, 4)
    wsQ.Range("E18").Value = Round(buyUSD, 2)
    wsQ.Range("G18").Value = Round(sellUSD, 2)
    wsQ.Range("I18").Value = Round(buyUSD - sellUSD, 2)
    
    Application.ScreenUpdating = True
End Sub
"""

print("Injecting VBA via win32com...")
import win32com.client, os, pythoncom

pythoncom.CoInitialize()
xl = win32com.client.Dispatch("Excel.Application")
xl.Visible = False
xl.DisplayAlerts = False

try:
    wb_com = xl.Workbooks.Open(xlsx_path)

    # 把 AllTrades 分頁名稱加進去（因為 VBA 需要用 ASCII 名稱）
    # 先把「總表」改名為 AllTrades
    for sh in wb_com.Sheets:
        if sh.Name == "\u7e3d\u8868":  # 總表
            sh.Name = "AllTrades"
            break

    # 注入 Module1（SearchUser 函數）
    vba_project = wb_com.VBProject
    mod = vba_project.VBComponents.Add(1)  # 1 = vbext_ct_StdModule
    mod.Name = "SearchModule"
    mod.CodeModule.AddFromString(VBA_MODULE_CODE)

    # 注入 查詢 sheet 的 Worksheet_Change 事件
    # 先找到「查詢」sheet 的 CodeName
    query_codename = None
    query_sheet = None
    for sh in wb_com.Sheets:
        if sh.Name == "\u67e5\u8a62":  # 查詢
            query_codename = sh.CodeName
            query_sheet = sh
            break
    print(f"  Query sheet CodeName: {query_codename}")
    if query_codename:
        for comp in vba_project.VBComponents:
            if comp.Name == query_codename:
                comp.CodeModule.AddFromString(VBA_SHEET_CODE)
                print(f"  Worksheet_Change injected into {query_codename}")
                break

    # 在查詢分頁 B15 右邊新增「清除」按鈕（C15 位置）
    if query_sheet is not None:
        c15 = query_sheet.Range("C15")
        btn = query_sheet.Buttons().Add(
            c15.Left + 4,
            c15.Top + 2,
            c15.Width * 1.6,
            c15.Height - 4,
        )
        btn.OnAction = "ClearSearch"
        btn.Characters.Text = "清除"
        btn.Font.Bold = True
        btn.Font.Size = 11
        print("  清除按鈕已新增")

    # 另存為 xlsm
    if os.path.exists(xlsm_path):
        try:
            os.remove(xlsm_path)
        except PermissionError:
            import time as _time

            print("  警告：xlsm 被佔用，等待 3 秒後重試...")
            _time.sleep(3)
            try:
                os.remove(xlsm_path)
            except PermissionError:
                # 改用新檔名避免衝突
                xlsm_path = xlsm_path.replace(".xlsm", "_new.xlsm")
                print(f"  改存為: {xlsm_path}")
    wb_com.SaveAs(xlsm_path, FileFormat=52)  # 52 = xlOpenXMLWorkbookMacroEnabled
    wb_com.Close(False)
    print(f"Saved xlsm: {xlsm_path}")
finally:
    xl.Quit()
    pythoncom.CoUninitialize()

# 清理暫存檔
if os.path.exists(xlsx_path):
    os.remove(xlsx_path)

print(f"\nDone! Output: {xlsm_path}")
print(f"Total rows: {total_rows}")
