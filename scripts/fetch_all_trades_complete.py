"""
fetch_all_trades_complete.py
完整版：Polymarket API (964筆) + Blockscout鏈上補錄
─────────────────────────────────────────────────────
資料來源：
  1. data-api.polymarket.com/trades?market=  ← 主要來源
  2. Blockscout ERC-1155 token transfers       ← 補漏用
     - 真實 matchOrders（單一市場方向性交易）
     - 排除 split/merge（price≈1.0 且雙向outcome）
     - 排除已在API中的 tx

Excel結構：
  總表 / 2025-12 / 2026-01 / ... / 2026-04 / 鏈上補錄 / 統計摘要
"""

import requests, time, json
from datetime import datetime, timezone, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TZ8 = timezone(timedelta(hours=8))

# ── 常數 ────────────────────────────────────────────────────────────────────
MARKETS = {
    "KMT": "0xc0f076bc4d90a44df34a729277e9d1f294f0cb60d2c3b1b3800908b1e15b923b",
    "DPP": "0xea3b1d0099085f43a3098b3e1fbcbe62284ce1bda99384b3d46b82ff202ac016",
    "TPP": "0x68fc8d466ddc10a1ae37b52642f36b93e413cf98ba5fe3947c242a9a727b2e94",
}

TOKEN_MAP = {
    "85632914518786177256583369552125280053108667306405854845853340618248288927460": (
        "KMT",
        "Yes",
    ),
    "4696955573632845407532815267539406678302911508204032661527405293140196109387": (
        "KMT",
        "No",
    ),
    "13628189982642424912108657221169198338993179248246381972030640500448717195916": (
        "DPP",
        "Yes",
    ),
    "91004506882941445266754771479824617369805789899332711132070603219216406556613": (
        "DPP",
        "No",
    ),
    "14999500579901383072635205035227864886528710236540822730141548371372688859422": (
        "TPP",
        "Yes",
    ),
    "16222840603445450947154718759167300491302153317593739623696847197718420087623": (
        "TPP",
        "No",
    ),
}

CTF = "0x4d97dcd97ec945f40cf65f87097ace5ea0476045"
NEG_EX = "0xc5d563a36ae78145c45a50134d48a1215220f80a"  # NegRiskCtfExchange
NEG_ADAPTER = "0xd91e80cf2e7be2e162c6513ced06f1dd0da35296"  # NegRisk adapter
ZERO = "0x0000000000000000000000000000000000000000"

KNOWN_CONTRACTS = {
    CTF,
    NEG_EX,
    NEG_ADAPTER,
    ZERO,
    "0x4bfb41d5b3570defd03c39a9a4d8de6bd8b8982e",  # UMA
    "0x3a3bd7bb9528e1ac2bc9d5ee5ec46c4b5ceaede1",  # LP pool?
    "0x2afd0dfad042ef1a16b07a603f28d20ea3b65f80",  # maker
}

PARTY_COLORS = {
    "KMT": ("002868", "FFFFFF"),
    "DPP": ("1B9431", "FFFFFF"),
    "TPP": ("28C8C8", "000000"),
}

HEADERS = [
    "時間 (UTC+8)",
    "政黨",
    "方向",
    "標的 (Yes/No)",
    "數量 (Shares)",
    "價格 ($)",
    "總金額 ($)",
    "交易者名稱",
    "錢包地址",
    "交易 Hash",
    "資料來源",
]
COL_WIDTHS = [22, 6, 6, 14, 16, 10, 12, 20, 44, 68, 12]


# ── API helpers ──────────────────────────────────────────────────────────────


def fetch_poly_trades(condition_id, party):
    trades = []
    offset = 0
    print(f"  Polymarket {party}...", end="", flush=True)
    while True:
        r = requests.get(
            f"https://data-api.polymarket.com/trades?market={condition_id}&limit=500&offset={offset}",
            timeout=30,
        )
        batch = r.json()
        if not batch:
            break
        trades.extend(batch)
        offset += 500
        time.sleep(0.25)
        if len(batch) < 500:
            break
    print(f" {len(trades)}筆")
    return trades


def fetch_blockscout_transfers(token_id):
    url = f"https://polygon.blockscout.com/api/v2/tokens/{CTF.capitalize().replace('x', 'X', 1)}/instances/{token_id}/transfers"
    url = f"https://polygon.blockscout.com/api/v2/tokens/0x4D97DCd97eC945f40cF65F87097ACe5EA0476045/instances/{token_id}/transfers"
    items = []
    page_params = None
    while True:
        r = requests.get(url, params=page_params, timeout=30)
        if r.status_code != 200:
            break
        d = r.json()
        items.extend(d.get("items", []))
        nxt = d.get("next_page_params")
        if not nxt:
            break
        page_params = nxt
        time.sleep(0.3)
    return items


def get_tx_token_transfers(tx_hash):
    r = requests.get(
        f"https://polygon.blockscout.com/api/v2/transactions/{tx_hash}/token-transfers",
        timeout=20,
    )
    if r.status_code == 200:
        return r.json().get("items", [])
    return []


def get_poly_profile(wallet):
    """Get Polymarket name/pseudonym for a wallet."""
    try:
        r = requests.get(
            f"https://data-api.polymarket.com/profiles?addresses={wallet}", timeout=10
        )
        if r.status_code == 200:
            d = r.json()
            if isinstance(d, list) and d:
                p = d[0]
                return p.get("name", "") or p.get("pseudonym", "")
    except:
        pass
    return ""


def ts_to_dt8(ts):
    return datetime.fromtimestamp(ts, tz=TZ8).strftime("%Y-%m-%d %H:%M:%S")


def month_key(ts):
    return datetime.fromtimestamp(ts, tz=TZ8).strftime("%Y-%m")


# ── Step 1: Polymarket API ───────────────────────────────────────────────────
print("=" * 55)
print("Step 1: Polymarket trades API")
print("=" * 55)
poly_by_party = {}
poly_hashes = set()

for party, cid in MARKETS.items():
    trades = fetch_poly_trades(cid, party)
    trades.sort(key=lambda x: x.get("timestamp", 0))
    for t in trades:
        h = t.get("transactionHash", "")
        if h:
            poly_hashes.add(h.lower())
    poly_by_party[party] = trades

total_poly = sum(len(v) for v in poly_by_party.values())
print(f"  合計: {total_poly}筆, unique hashes: {len(poly_hashes)}")


# ── Step 2: Blockscout on-chain ──────────────────────────────────────────────
print()
print("=" * 55)
print("Step 2: Blockscout token transfers (補漏)")
print("=" * 55)

# Collect all tx_hash → {party, outcome, transfers, users}
onchain_txs = {}  # tx_hash (lower) → dict

for tid, (party, outcome) in TOKEN_MAP.items():
    print(f"  {party}-{outcome}...", end="", flush=True)
    items = fetch_blockscout_transfers(tid)
    print(f" {len(items)} transfers")

    for item in items:
        tx_hash = item.get("transaction_hash", "").lower()
        if not tx_hash or tx_hash in poly_hashes:
            continue  # already in Polymarket API

        from_addr = item.get("from", {}).get("hash", "").lower()
        to_addr = item.get("to", {}).get("hash", "").lower()
        amount_raw = item.get("total", {}).get("value", "0")
        try:
            amount = int(amount_raw) / 1_000_000
        except:
            amount = 0
        ts_str = item.get("timestamp", "")
        ttype = item.get("type", "")

        if tx_hash not in onchain_txs:
            onchain_txs[tx_hash] = {
                "tx_hash": tx_hash,
                "timestamp_str": ts_str,
                "transfers": [],
                "parties": set(),
                "outcomes": set(),
            }

        onchain_txs[tx_hash]["transfers"].append(
            {
                "party": party,
                "outcome": outcome,
                "from": from_addr,
                "to": to_addr,
                "amount": amount,
                "type": ttype,
            }
        )
        onchain_txs[tx_hash]["parties"].add(party)
        onchain_txs[tx_hash]["outcomes"].add(outcome)

print(f"  未在API中的唯一鏈上tx: {len(onchain_txs)}")


# ── Step 3: Classify & enrich missing txs ────────────────────────────────────
print()
print("=" * 55)
print("Step 3: 分類 & 補充鏈上交易資訊")
print("=" * 55)

supplemental_trades = []  # list of trade dicts compatible with Excel rows


def classify_and_enrich(tx_hash, tx_info):
    transfers = tx_info["transfers"]
    parties = tx_info["parties"]
    outcomes = tx_info["outcomes"]
    ts_str = tx_info["timestamp_str"]

    # If both Yes and No outcomes are touched at equal amounts → split/merge, skip
    yes_amounts = [
        tr["amount"] for tr in transfers if tr["outcome"] == "Yes" and tr["amount"] > 0
    ]
    no_amounts = [
        tr["amount"] for tr in transfers if tr["outcome"] == "No" and tr["amount"] > 0
    ]
    if yes_amounts and no_amounts and abs(max(yes_amounts) - max(no_amounts)) < 0.01:
        return None  # split/merge, not directional

    # Determine outcome direction
    outcome = (
        "Yes"
        if (max(yes_amounts) if yes_amounts else 0)
        >= (max(no_amounts) if no_amounts else 0)
        else "No"
    )
    shares = max(yes_amounts + no_amounts) if (yes_amounts + no_amounts) else 0
    if shares < 0.001:
        return None

    # Determine side: minting → BUY, burning → SELL
    sides = set()
    for tr in transfers:
        if tr["type"] == "token_minting":
            sides.add("BUY")
        elif tr["type"] == "token_burning":
            sides.add("SELL")
    if not sides:
        return None  # can't determine
    side = "BUY" if "BUY" in sides else "SELL"

    # Primary party (if single, use it; if multi → pick most-represented)
    if len(parties) == 1:
        party = list(parties)[0]
    else:
        # Count by outcome amount
        party_amounts = defaultdict(float)
        for tr in transfers:
            party_amounts[tr["party"]] += tr["amount"]
        party = max(party_amounts, key=party_amounts.get)

    # Get USDC from full tx token-transfers
    tx_trs = get_tx_token_transfers(tx_hash)
    time.sleep(0.25)

    # Find user address
    all_addrs = set()
    for tr in transfers:
        for a in [tr["from"], tr["to"]]:
            if a and a not in KNOWN_CONTRACTS:
                all_addrs.add(a)
    user_addr = next(iter(all_addrs), "")

    # USDC amount: outgoing from user (BUY) or incoming to user (SELL)
    usdc_amount = 0.0
    for tr in tx_trs:
        sym = tr.get("token", {}).get("symbol", "") or ""
        if "USDC" not in sym.upper() and "WCOL" not in sym.upper():
            continue
        decimals = int(tr.get("total", {}).get("decimals", 6) or 6)
        try:
            val = int(tr.get("total", {}).get("value", 0)) / (10**decimals)
        except:
            val = 0
        fr = tr.get("from", {}).get("hash", "").lower()
        to = tr.get("to", {}).get("hash", "").lower()
        if side == "BUY" and user_addr and fr == user_addr:
            usdc_amount = max(usdc_amount, val)
        elif side == "SELL" and user_addr and to == user_addr:
            usdc_amount = max(usdc_amount, val)

    if usdc_amount == 0:
        # fallback: any WCOL minting = cost for BUY, WCOL burning = proceeds for SELL
        for tr in tx_trs:
            sym = tr.get("token", {}).get("symbol", "") or ""
            if "WCOL" in sym.upper() or "USDC" in sym.upper():
                try:
                    decimals = int(tr.get("total", {}).get("decimals", 6) or 6)
                    val = int(tr.get("total", {}).get("value", 0)) / (10**decimals)
                    if val > usdc_amount:
                        usdc_amount = val
                except:
                    pass

    price = round(usdc_amount / shares, 4) if shares > 0 and usdc_amount > 0 else 0

    # Skip if price ≈ 1.0 (split operation, not a trade)
    if price > 0.97:
        return None

    # Get Polymarket profile
    name = get_poly_profile(user_addr) if user_addr else ""
    time.sleep(0.2)

    # Parse timestamp
    try:
        dt_utc = datetime.fromisoformat(ts_str.replace("Z", "+00:00"))
        ts_epoch = int(dt_utc.timestamp())
    except:
        ts_epoch = 0

    return {
        "timestamp": ts_epoch,
        "party": party,
        "side": side,
        "outcome": outcome,
        "size": round(shares, 6),
        "price": price,
        "total": round(usdc_amount, 4),
        "name": name,
        "proxyWallet": user_addr,
        "transactionHash": tx_hash,
        "source": "鏈上補錄",
    }


total_onchain = len(onchain_txs)
added = 0
skipped_split = 0
skipped_unclear = 0

for i, (tx_hash, tx_info) in enumerate(onchain_txs.items()):
    if (i + 1) % 10 == 0:
        print(f"  [{i + 1}/{total_onchain}] added={added}...", flush=True)
    result = classify_and_enrich(tx_hash, tx_info)
    if result is None:
        skipped_split += 1
    else:
        supplemental_trades.append(result)
        added += 1

print(f"  鏈上補錄: {added}筆 | split/merge略過: {skipped_split}筆")


# ── Step 4: Merge all data ────────────────────────────────────────────────────
print()
print("=" * 55)
print("Step 4: 合併資料")
print("=" * 55)

# Convert poly trades to unified format
all_flat = []

for party, trades in poly_by_party.items():
    for t in trades:
        all_flat.append(
            {
                "timestamp": t.get("timestamp", 0),
                "party": party,
                "side": t.get("side", ""),
                "outcome": t.get("outcome", ""),
                "size": round(float(t.get("size", 0)), 6),
                "price": round(float(t.get("price", 0)), 4),
                "total": round(float(t.get("size", 0)) * float(t.get("price", 0)), 4),
                "name": t.get("name", "") or t.get("pseudonym", "") or "",
                "proxyWallet": t.get("proxyWallet", ""),
                "transactionHash": t.get("transactionHash", ""),
                "source": "Polymarket API",
            }
        )

# Add supplemental on-chain trades
all_flat.extend(supplemental_trades)

# Sort by timestamp
all_flat.sort(key=lambda x: x["timestamp"])

print(f"  Polymarket API: {total_poly}筆")
print(f"  鏈上補錄:       {added}筆")
print(f"  合計:           {len(all_flat)}筆")

# Group by month
monthly = defaultdict(list)
for t in all_flat:
    if t["timestamp"] > 0:
        mk = month_key(t["timestamp"])
        monthly[mk].append(t)

monthly_keys = sorted(monthly.keys())
print(f"  月份: {monthly_keys[0]} ～ {monthly_keys[-1]}")
for mk in monthly_keys:
    print(f"    {mk}: {len(monthly[mk])}筆")


# ── Step 5: Build Excel ───────────────────────────────────────────────────────
print()
print("=" * 55)
print("Step 5: 建立 Excel")
print("=" * 55)


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def write_header(ws, border):
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(fill_type="solid", fgColor="1F3864")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


def write_rows(ws, trades, border):
    for i, t in enumerate(trades):
        row_idx = i + 2
        party = t["party"]
        side = t["side"]
        bg_color, fg_color = PARTY_COLORS.get(party, ("888888", "FFFFFF"))
        alt_bg = "EBF5FB" if i % 2 == 0 else "FFFFFF"
        source = t.get("source", "")

        # Light orange tint for on-chain supplemental
        if source == "鏈上補錄":
            alt_bg = "FFF3E0" if i % 2 == 0 else "FFFDE7"

        row_data = [
            ts_to_dt8(t["timestamp"]) if t["timestamp"] else "",
            party,
            side,
            t["outcome"],
            t["size"],
            t["price"],
            t["total"],
            t["name"],
            t["proxyWallet"],
            t["transactionHash"],
            source,
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            if col_idx == 2:  # 政黨
                cell.fill = PatternFill(fill_type="solid", fgColor=bg_color)
                cell.font = Font(bold=True, color=fg_color, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 3:  # BUY/SELL
                color = "006400" if side == "BUY" else "8B0000"
                cell.font = Font(color=color, bold=True, size=10)
                cell.fill = PatternFill(fill_type="solid", fgColor=alt_bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 4:  # Yes/No
                cell.font = Font(size=10)
                cell.fill = PatternFill(fill_type="solid", fgColor=alt_bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 11:  # 資料來源
                cell.font = Font(
                    size=9,
                    italic=(source == "鏈上補錄"),
                    color="B8860B" if source == "鏈上補錄" else "444444",
                )
                cell.fill = PatternFill(fill_type="solid", fgColor=alt_bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = Font(size=10)
                cell.fill = PatternFill(fill_type="solid", fgColor=alt_bg)


def write_summary_sheet(wb, all_trades, monthly_keys, monthly, border):
    ws = wb.create_sheet(title="統計摘要")
    # Party stats
    stat_headers = [
        "政黨",
        "總筆數",
        "BUY",
        "SELL",
        "Yes總量",
        "No總量",
        "總金額($)",
        "平均價格($)",
        "API筆數",
        "鏈上補錄筆數",
    ]
    col_widths_s = [8, 8, 8, 8, 12, 12, 14, 14, 10, 12]
    for col_idx, h in enumerate(stat_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(fill_type="solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths_s[
            col_idx - 1
        ]
    ws.row_dimensions[1].height = 28

    by_party = defaultdict(list)
    for t in all_trades:
        by_party[t["party"]].append(t)

    for row_idx, party in enumerate(["KMT", "DPP", "TPP"], 2):
        trades = by_party[party]
        buys = [t for t in trades if t["side"] == "BUY"]
        sells = [t for t in trades if t["side"] == "SELL"]
        yes_v = sum(t["size"] for t in trades if t["outcome"] == "Yes")
        no_v = sum(t["size"] for t in trades if t["outcome"] == "No")
        total = sum(t["total"] for t in trades)
        avg_p = sum(t["price"] for t in trades) / len(trades) if trades else 0
        api_n = sum(1 for t in trades if t.get("source") == "Polymarket API")
        onc_n = sum(1 for t in trades if t.get("source") == "鏈上補錄")

        row_data = [
            party,
            len(trades),
            len(buys),
            len(sells),
            round(yes_v, 2),
            round(no_v, 2),
            round(total, 2),
            round(avg_p, 4),
            api_n,
            onc_n,
        ]
        bg, fg = PARTY_COLORS[party]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            cell.fill = PatternFill(
                fill_type="solid", fgColor=bg if col_idx == 1 else "F2F2F2"
            )
            cell.font = Font(
                bold=(col_idx == 1), color=fg if col_idx == 1 else "000000", size=11
            )

    # Monthly summary
    ws.cell(row=6, column=1, value="--- 月份分布 ---").font = Font(bold=True, size=11)
    monthly_h = ["月份", "總筆數", "KMT", "DPP", "TPP", "鏈上補錄"]
    for col_idx, h in enumerate(monthly_h, 1):
        cell = ws.cell(row=7, column=col_idx, value=h)
        cell.fill = PatternFill(fill_type="solid", fgColor="2E4057")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r_off, mk in enumerate(monthly_keys, 8):
        trades_m = monthly[mk]
        kmt_n = sum(1 for t in trades_m if t["party"] == "KMT")
        dpp_n = sum(1 for t in trades_m if t["party"] == "DPP")
        tpp_n = sum(1 for t in trades_m if t["party"] == "TPP")
        onc_n = sum(1 for t in trades_m if t.get("source") == "鏈上補錄")
        row_data_m = [mk, len(trades_m), kmt_n, dpp_n, tpp_n, onc_n]
        for col_idx, val in enumerate(row_data_m, 1):
            cell = ws.cell(row=r_off, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            cell.fill = PatternFill(fill_type="solid", fgColor="F9F9F9")
            cell.font = Font(size=10)


wb = Workbook()
border = thin_border()

# 總表
ws_all = wb.active
ws_all.title = "總表"
write_header(ws_all, border)
write_rows(ws_all, all_flat, border)
print(f"  總表: {len(all_flat)}筆 ✓")

# 月份工作表
for mk in monthly_keys:
    ws_m = wb.create_sheet(title=mk)
    write_header(ws_m, border)
    write_rows(ws_m, monthly[mk], border)
    print(f"  {mk}: {len(monthly[mk])}筆 ✓")

# 鏈上補錄工作表（若有）
if supplemental_trades:
    ws_oc = wb.create_sheet(title="鏈上補錄")
    write_header(ws_oc, border)
    write_rows(ws_oc, sorted(supplemental_trades, key=lambda x: x["timestamp"]), border)
    print(f"  鏈上補錄: {len(supplemental_trades)}筆 ✓")

# 統計摘要
write_summary_sheet(wb, all_flat, monthly_keys, monthly, border)
print("  統計摘要 ✓")

output = "E:\\polymarket選舉賭博\\2026台灣地方選舉_全部交易.xlsx"
wb.save(output)
print(f"\nExcel 儲存至: {output}")
print(f"總計: {len(all_flat)} 筆 (API={total_poly}, 鏈上補錄={added})")
